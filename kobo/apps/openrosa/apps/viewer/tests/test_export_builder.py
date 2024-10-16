# coding: utf-8
import datetime
import os

from django.conf import settings
from django.core.files.temp import NamedTemporaryFile
from openpyxl import load_workbook
from pyxform.builder import create_survey_from_xls

from kobo.apps.openrosa.apps.main.tests.test_base import TestBase
from kobo.apps.openrosa.apps.viewer.tests.export_helpers import viewer_fixture_path
from kobo.apps.openrosa.libs.utils.export_tools import (
    dict_to_joined_export,
    ExportBuilder,
)
from kpi.utils.mongo_helper import MongoHelper


def _logger_fixture_path(*args):
    return os.path.join(settings.OPENROSA_APP_DIR, 'apps', 'logger',
                        'tests', 'fixtures', *args)


class TestExportBuilder(TestBase):
    data = [
        {
            'name': 'Abe',
            'age': 35,
            'tel/telLg==office': '020123456',
            'children':
            [
                {
                    'children/name': 'Mike',
                    'children/age': 5,
                    'children/fav_colors': 'red blue',
                    'children/iceLg==creams': 'vanilla chocolate',
                    'children/cartoons':
                    [
                        {
                            'children/cartoons/name': 'Tom & Jerry',
                            'children/cartoons/why': 'Tom is silly',
                        },
                        {
                            'children/cartoons/name': 'Flinstones',
                            'children/cartoons/why': "I like bam bam\u0107"
                            # throw in a unicode character
                        }
                    ]
                },
                {
                    'children/name': 'John',
                    'children/age': 2,
                    'children/cartoons': []
                },
                {
                    'children/name': 'Imora',
                    'children/age': 3,
                    'children/cartoons':
                    [
                        {
                            'children/cartoons/name': 'Shrek',
                            'children/cartoons/why': 'He\'s so funny'
                        },
                        {
                            'children/cartoons/name': 'Dexter\'s Lab',
                            'children/cartoons/why': 'He thinks hes smart',
                            'children/cartoons/characters':
                            [
                                {
                                    'children/cartoons/characters/name':
                                    'Dee Dee',
                                    'children/cartoons/characters/good_or_evi'
                                    'l': 'good'
                                },
                                {
                                    'children/cartoons/characters/name':
                                    'Dexter',
                                    'children/cartoons/characters/good_or_evi'
                                    'l': 'evil'
                                },
                            ]
                        }
                    ]
                }
            ]
        },
        {
            # blank data just to be sure
            'children': []
        }
    ]
    long_survey_data = [
        {
            'name': 'Abe',
            'age': 35,
            'childrens_survey_with_a_very_lo':
            [
                {
                    'childrens_survey_with_a_very_lo/name': 'Mike',
                    'childrens_survey_with_a_very_lo/age': 5,
                    'childrens_survey_with_a_very_lo/fav_colors': 'red blue',
                    'childrens_survey_with_a_very_lo/cartoons':
                    [
                        {
                            'childrens_survey_with_a_very_lo/cartoons/name':
                            'Tom & Jerry',
                            'childrens_survey_with_a_very_lo/cartoons/why':
                            'Tom is silly',
                        },
                        {
                            'childrens_survey_with_a_very_lo/cartoons/name':
                            'Flinstones',
                            'childrens_survey_with_a_very_lo/cartoons/why':
                            "I like bam bam\u0107"
                            # throw in a unicode character
                        }
                    ]
                },
                {
                    'childrens_survey_with_a_very_lo/name': 'John',
                    'childrens_survey_with_a_very_lo/age': 2,
                    'childrens_survey_with_a_very_lo/cartoons': []
                },
                {
                    'childrens_survey_with_a_very_lo/name': 'Imora',
                    'childrens_survey_with_a_very_lo/age': 3,
                    'childrens_survey_with_a_very_lo/cartoons':
                    [
                        {
                            'childrens_survey_with_a_very_lo/cartoons/name':
                            'Shrek',
                            'childrens_survey_with_a_very_lo/cartoons/why':
                            'He\'s so funny'
                        },
                        {
                            'childrens_survey_with_a_very_lo/cartoons/name':
                            'Dexter\'s Lab',
                            'childrens_survey_with_a_very_lo/cartoons/why':
                            'He thinks hes smart',
                            'childrens_survey_with_a_very_lo/cartoons/characte'
                            'rs':
                            [
                                {
                                    'childrens_survey_with_a_very_lo/cartoons/'
                                    'characters/name': 'Dee Dee',
                                    'children/cartoons/characters/good_or_evi'
                                    'l': 'good'
                                },
                                {
                                    'childrens_survey_with_a_very_lo/cartoons/'
                                    'characters/name': 'Dexter',
                                    'children/cartoons/characters/good_or_evi'
                                    'l': 'evil'
                                },
                            ]
                        }
                    ]
                }
            ]
        }
    ]
    data_utf8 = [
        {
            'name': 'Abe',
            'age': 35,
            'tel/telLg==office': '020123456',
            'childrenLg==info':
            [
                {
                    'childrenLg==info/nameLg==first': 'Mike',
                    'childrenLg==info/age': 5,
                    'childrenLg==info/fav_colors': 'red\u2019s blue\u2019s',
                    'childrenLg==info/ice_creams': 'vanilla chocolate',
                    'childrenLg==info/cartoons':
                    [
                        {
                            'childrenLg==info/cartoons/name': 'Tom & Jerry',
                            'childrenLg==info/cartoons/why': 'Tom is silly',
                        },
                        {
                            'childrenLg==info/cartoons/name': 'Flinstones',
                            'childrenLg==info/cartoons/why':
                            "I like bam bam\u0107"
                            # throw in a unicode character
                        }
                    ]
                }
            ]
        }
    ]

    def _create_childrens_survey(self):
        return create_survey_from_xls(_logger_fixture_path(
            'childrens_survey.xls'),
            default_name='childrens_survey')

    def test_build_sections_from_survey(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        # test that we generate the proper sections
        expected_sections = [
            survey.name, 'children', 'children/cartoons',
            'children/cartoons/characters']
        self.assertEqual(
            expected_sections, [s['name'] for s in export_builder.sections])
        # main section should have split geolocations
        expected_element_names = [
            'name', 'age', 'geo/geolocation', 'geo/_geolocation_longitude',
            'geo/_geolocation_latitude', 'geo/_geolocation_altitude',
            'geo/_geolocation_precision', 'tel/tel.office', 'tel/tel.mobile',
            'meta/instanceID']
        section = export_builder.section_by_name(survey.name)
        element_names = [element['xpath'] for element in section['elements']]
        # fav_colors should have its choices split
        self.assertEqual(
            sorted(expected_element_names), sorted(element_names))

        expected_element_names = [
            'children/name', 'children/age', 'children/fav_colors',
            'children/fav_colors/red', 'children/fav_colors/blue',
            'children/fav_colors/pink', 'children/ice.creams',
            'children/ice.creams/vanilla', 'children/ice.creams/strawberry',
            'children/ice.creams/chocolate']
        section = export_builder.section_by_name('children')
        element_names = [element['xpath'] for element in section['elements']]
        self.assertEqual(
            sorted(expected_element_names), sorted(element_names))

        expected_element_names = [
            'children/cartoons/name', 'children/cartoons/why']
        section = export_builder.section_by_name('children/cartoons')
        element_names = [element['xpath'] for element in section['elements']]

        self.assertEqual(
            sorted(expected_element_names), sorted(element_names))

        expected_element_names = [
            'children/cartoons/characters/name',
            'children/cartoons/characters/good_or_evil']
        section = \
            export_builder.section_by_name('children/cartoons/characters')
        element_names = [element['xpath'] for element in section['elements']]
        self.assertEqual(
            sorted(expected_element_names), sorted(element_names))

    def test_decode_mongo_encoded_section_names(self):
        data = {
            'main_section': [1, 2, 3, 4],
            'sectionLg==1/info': [1, 2, 3, 4],
            'sectionLg==2/info': [1, 2, 3, 4],
        }
        result = ExportBuilder.decode_mongo_encoded_section_names(data)
        expected_result = {
            'main_section': [1, 2, 3, 4],
            'section.1/info': [1, 2, 3, 4],
            'section.2/info': [1, 2, 3, 4],
        }
        self.assertEqual(result, expected_result)

    def test_generation_of_multi_selects_works(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        expected_select_multiples =\
            {
                'children':
                {
                    'children/fav_colors':
                    [
                        'children/fav_colors/red', 'children/fav_colors/blue',
                        'children/fav_colors/pink'
                    ],
                    'children/ice.creams':
                    [
                        'children/ice.creams/vanilla',
                        'children/ice.creams/strawberry',
                        'children/ice.creams/chocolate'
                    ]
                }
            }
        select_multiples = export_builder.select_multiples
        self.assertTrue('children' in select_multiples)
        self.assertTrue('children/fav_colors' in select_multiples['children'])
        self.assertTrue('children/ice.creams' in select_multiples['children'])
        self.assertEqual(
            sorted(select_multiples['children']['children/fav_colors']),
            sorted(
                expected_select_multiples['children']['children/fav_colors']))
        self.assertEqual(
            sorted(select_multiples['children']['children/ice.creams']),
            sorted(
                expected_select_multiples['children']['children/ice.creams']))

    def test_split_select_multiples_works(self):
        select_multiples =\
            {
                'children/fav_colors': [
                    'children/fav_colors/red', 'children/fav_colors/blue',
                    'children/fav_colors/pink']
            }
        row = \
            {
                'children/name': 'Mike',
                'children/age': 5,
                'children/fav_colors': 'red blue'
            }
        new_row = ExportBuilder.split_select_multiples(
            row, select_multiples)
        expected_row = \
            {
                'children/name': 'Mike',
                'children/age': 5,
                'children/fav_colors': 'red blue',
                'children/fav_colors/red': True,
                'children/fav_colors/blue': True,
                'children/fav_colors/pink': False
            }
        self.assertEqual(new_row, expected_row)
        row = \
            {
                'children/name': 'Mike',
                'children/age': 5,
            }
        new_row = ExportBuilder.split_select_multiples(
            row, select_multiples)
        expected_row = \
            {
                'children/name': 'Mike',
                'children/age': 5,
                'children/fav_colors/red': None,
                'children/fav_colors/blue': None,
                'children/fav_colors/pink': None
            }
        self.assertEqual(new_row, expected_row)

    def test_split_select_multiples_works_when_data_is_blank(self):
        select_multiples =\
            {
                'children/fav_colors': [
                    'children/fav_colors/red', 'children/fav_colors/blue',
                    'children/fav_colors/pink']
            }
        row = \
            {
                'children/name': 'Mike',
                'children/age': 5,
                'children/fav_colors': ''
            }
        new_row = ExportBuilder.split_select_multiples(
            row, select_multiples)
        expected_row = \
            {
                'children/name': 'Mike',
                'children/age': 5,
                'children/fav_colors': '',
                'children/fav_colors/red': None,
                'children/fav_colors/blue': None,
                'children/fav_colors/pink': None
            }
        self.assertEqual(new_row, expected_row)

    def test_generation_of_gps_fields_works(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        expected_gps_fields =\
            {
                'childrens_survey':
                {
                    'geo/geolocation':
                    [
                        'geo/_geolocation_latitude',
                        'geo/_geolocation_longitude',
                        'geo/_geolocation_altitude',
                        'geo/_geolocation_precision'
                    ]
                }
            }
        gps_fields = export_builder.gps_fields
        self.assertTrue('childrens_survey' in gps_fields)
        self.assertEqual(
            sorted(gps_fields['childrens_survey']),
            sorted(expected_gps_fields['childrens_survey']))

    def test_split_gps_components_works(self):
        gps_fields =\
            {
                'geo/geolocation':
                [
                    'geo/_geolocation_latitude', 'geo/_geolocation_longitude',
                    'geo/_geolocation_altitude', 'geo/_geolocation_precision'
                ]
            }
        row = \
            {
                'geo/geolocation': '1.0 36.1 2000 20',
            }
        new_row = ExportBuilder.split_gps_components(
            row, gps_fields)
        expected_row = \
            {
                'geo/geolocation': '1.0 36.1 2000 20',
                'geo/_geolocation_latitude': '1.0',
                'geo/_geolocation_longitude': '36.1',
                'geo/_geolocation_altitude': '2000',
                'geo/_geolocation_precision': '20'
            }
        self.assertEqual(new_row, expected_row)

    def test_split_gps_components_works_when_gps_data_is_blank(self):
        gps_fields =\
            {
                'geo/geolocation':
                [
                    'geo/_geolocation_latitude', 'geo/_geolocation_longitude',
                    'geo/_geolocation_altitude', 'geo/_geolocation_precision'
                ]
            }
        row = \
            {
                'geo/geolocation': '',
            }
        new_row = ExportBuilder.split_gps_components(
            row, gps_fields)
        expected_row = \
            {
                'geo/geolocation': '',
            }
        self.assertEqual(new_row, expected_row)

    def test_generation_of_mongo_encoded_fields_works(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        expected_encoded_fields =\
            {
                'childrens_survey':
                {
                    'tel/tel.office': 'tel/{0}'.format(
                        MongoHelper.encode('tel.office')),
                    'tel/tel.mobile': 'tel/{0}'.format(
                        MongoHelper.encode('tel.mobile')),
                }
            }
        encoded_fields = export_builder.encoded_fields
        self.assertTrue('childrens_survey' in encoded_fields)
        self.assertEqual(
            encoded_fields['childrens_survey'],
            expected_encoded_fields['childrens_survey'])

    def test_decode_fields_names_encoded_for_mongo(self):
        encoded_fields = \
            {
                'tel/tel.office': 'tel/{0}'.format(
                    MongoHelper.encode('tel.office'))
            }
        row = \
            {
                'name': 'Abe',
                'age': 35,
                'tel/{0}'.format(
                    MongoHelper.encode('tel.office')): '123-456-789'
            }
        new_row = ExportBuilder.decode_mongo_encoded_fields(
            row, encoded_fields)
        expected_row = \
            {
                'name': 'Abe',
                'age': 35,
                'tel/tel.office': '123-456-789'
            }
        self.assertEqual(new_row, expected_row)

    def test_generate_field_title(self):
        field_name = ExportBuilder.format_field_title("child/age", ".")
        expected_field_name = "child.age"
        self.assertEqual(field_name, expected_field_name)

    def test_delimiter_replacement_works_existing_fields(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.GROUP_DELIMITER = "."
        export_builder.set_survey(survey)
        expected_sections =\
            [
                {
                    'name': 'children',
                    'elements': [
                        {
                            'title': 'children.name',
                            'xpath': 'children/name'
                        }
                    ]
                }
            ]
        children_section = export_builder.section_by_name('children')
        self.assertEqual(
            children_section['elements'][0]['title'],
            expected_sections[0]['elements'][0]['title'])

    def test_delimiter_replacement_works_generated_multi_select_fields(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.GROUP_DELIMITER = "."
        export_builder.set_survey(survey)
        expected_section =\
            {
                'name': 'children',
                'elements': [
                    {
                        'title': 'children.fav_colors.red',
                        'xpath': 'children/fav_colors/red'
                    }
                ]
            }
        childrens_section = export_builder.section_by_name('children')
        match = [x for x in childrens_section['elements']
                 if x['xpath'] == expected_section['elements'][0]['xpath']][0]
        self.assertEqual(
            expected_section['elements'][0]['title'], match['title'])

    def test_delimiter_replacement_works_for_generated_gps_fields(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.GROUP_DELIMITER = "."
        export_builder.set_survey(survey)
        expected_section = \
            {
                'name': 'childrens_survey',
                'elements': [
                    {
                        'title': 'geo._geolocation_latitude',
                        'xpath': 'geo/_geolocation_latitude'
                    }
                ]
            }
        main_section = export_builder.section_by_name('childrens_survey')
        match = [x for x in main_section['elements'] if x['xpath'] == expected_section['elements'][0]['xpath']][0]
        self.assertEqual(
            expected_section['elements'][0]['title'], match['title'])

    def test_to_xlsx_export_works(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        xls_file = NamedTemporaryFile(suffix='.xlsx')
        filename = xls_file.name
        export_builder.to_xls_export(filename, self.data)
        xls_file.seek(0)
        wb = load_workbook(filename)
        # check that we have childrens_survey, children, children_cartoons
        # and children_cartoons_characters sheets
        expected_sheet_names = ['childrens_survey', 'children',
                                'children_cartoons',
                                'children_cartoons_characters']
        self.assertEqual(wb.sheetnames, expected_sheet_names)

        # check header columns
        main_sheet = wb['childrens_survey']
        expected_column_headers = [
            'name', 'age', 'geo/geolocation', 'geo/_geolocation_latitude',
            'geo/_geolocation_longitude', 'geo/_geolocation_altitude',
            'geo/_geolocation_precision', 'tel/tel.office',
            'tel/tel.mobile', '_id', 'meta/instanceID', '_uuid',
            '_submission_time', '_index', '_parent_index',
            '_parent_table_name', '_tags', '_notes']
        column_headers = [c[0].value for c in main_sheet.columns]
        self.assertEqual(sorted(column_headers),
                         sorted(expected_column_headers))

        childrens_sheet = wb['children']
        expected_column_headers = [
            'children/name', 'children/age', 'children/fav_colors',
            'children/fav_colors/red', 'children/fav_colors/blue',
            'children/fav_colors/pink', 'children/ice.creams',
            'children/ice.creams/vanilla', 'children/ice.creams/strawberry',
            'children/ice.creams/chocolate', '_id', '_uuid',
            '_submission_time', '_index', '_parent_index',
            '_parent_table_name', '_tags', '_notes']
        column_headers = [c[0].value for c in childrens_sheet.columns]
        self.assertEqual(sorted(column_headers),
                         sorted(expected_column_headers))

        cartoons_sheet = wb['children_cartoons']
        expected_column_headers = [
            'children/cartoons/name', 'children/cartoons/why', '_id',
            '_uuid', '_submission_time', '_index', '_parent_index',
            '_parent_table_name', '_tags', '_notes']
        column_headers = [c[0].value for c in cartoons_sheet.columns]
        self.assertEqual(sorted(column_headers),
                         sorted(expected_column_headers))

        characters_sheet = wb['children_cartoons_characters']
        expected_column_headers = [
            'children/cartoons/characters/name',
            'children/cartoons/characters/good_or_evil', '_id', '_uuid',
            '_submission_time', '_index', '_parent_index',
            '_parent_table_name', '_tags', '_notes']
        column_headers = [c[0].value for c in characters_sheet.columns]
        self.assertEqual(sorted(column_headers),
                         sorted(expected_column_headers))

        xls_file.close()

    def test_to_xlsx_export_respects_custom_field_delimiter(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.GROUP_DELIMITER = ExportBuilder.GROUP_DELIMITER_DOT
        export_builder.set_survey(survey)
        xls_file = NamedTemporaryFile(suffix='.xlsx')
        filename = xls_file.name
        export_builder.to_xls_export(filename, self.data)
        xls_file.seek(0)
        wb = load_workbook(filename)

        # check header columns
        main_sheet = wb['childrens_survey']
        expected_column_headers = [
            'name', 'age', 'geo.geolocation', 'geo._geolocation_latitude',
            'geo._geolocation_longitude', 'geo._geolocation_altitude',
            'geo._geolocation_precision', 'tel.tel.office',
            'tel.tel.mobile', '_id', 'meta.instanceID', '_uuid',
            '_submission_time', '_index', '_parent_index',
            '_parent_table_name', '_tags', '_notes']
        column_headers = [c[0].value for c in main_sheet.columns]
        self.assertEqual(sorted(column_headers),
                         sorted(expected_column_headers))
        xls_file.close()

    def test_get_valid_sheet_name_catches_duplicates(self):
        work_sheets = {'childrens_survey': "Worksheet"}
        desired_sheet_name = "childrens_survey"
        expected_sheet_name = "childrens_survey1"
        generated_sheet_name = ExportBuilder.get_valid_sheet_name(
            desired_sheet_name, work_sheets)
        self.assertEqual(generated_sheet_name, expected_sheet_name)

    def test_get_valid_sheet_name_catches_long_names(self):
        desired_sheet_name = "childrens_survey_with_a_very_long_name"
        expected_sheet_name = "childrens_survey_with_a_very_lo"
        generated_sheet_name = ExportBuilder.get_valid_sheet_name(
            desired_sheet_name, [])
        self.assertEqual(generated_sheet_name, expected_sheet_name)

    def test_get_valid_sheet_name_catches_long_duplicate_names(self):
        work_sheet_titles = ['childrens_survey_with_a_very_lo']
        desired_sheet_name = "childrens_survey_with_a_very_long_name"
        expected_sheet_name = "childrens_survey_with_a_very_l1"
        generated_sheet_name = ExportBuilder.get_valid_sheet_name(
            desired_sheet_name, work_sheet_titles)
        self.assertEqual(generated_sheet_name, expected_sheet_name)

    def test_to_xls_export_generates_valid_sheet_names(self):
        survey = create_survey_from_xls(_logger_fixture_path(
            'childrens_survey_with_a_very_long_name.xls'),
            default_name='childrens_survey_with_a_very_long_name')
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        xls_file = NamedTemporaryFile(suffix='.xlsx')
        filename = xls_file.name
        export_builder.to_xls_export(filename, self.data)
        xls_file.seek(0)
        wb = load_workbook(filename)
        # check that we have childrens_survey, children, children_cartoons
        # and children_cartoons_characters sheets
        expected_sheet_names = ['childrens_survey_with_a_very_lo',
                                'childrens_survey_with_a_very_l1',
                                'childrens_survey_with_a_very_l2',
                                'childrens_survey_with_a_very_l3']
        self.assertEqual(wb.sheetnames, expected_sheet_names)
        xls_file.close()

    def test_child_record_parent_table_is_updated_when_sheet_is_renamed(self):
        survey = create_survey_from_xls(_logger_fixture_path(
            'childrens_survey_with_a_very_long_name.xls'),
            default_name='childrens_survey_with_a_very_long_name.xls')
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        xls_file = NamedTemporaryFile(suffix='.xlsx')
        filename = xls_file.name
        export_builder.to_xls_export(filename, self.long_survey_data)
        xls_file.seek(0)
        wb = load_workbook(filename)

        # get the children's sheet
        ws1 = wb['childrens_survey_with_a_very_l1']
        # parent_table is in cell K2
        parent_table_name = ws1.cell(row=2, column=11).value
        expected_parent_table_name = 'childrens_survey_with_a_very_lo'
        self.assertEqual(parent_table_name, expected_parent_table_name)

        # get cartoons sheet
        ws2 = wb['childrens_survey_with_a_very_l2']
        parent_table_name = ws2.cell(row=2, column=7).value
        expected_parent_table_name = 'childrens_survey_with_a_very_l1'
        self.assertEqual(parent_table_name, expected_parent_table_name)
        xls_file.close()

    def test_type_conversion(self):
        submission_1 = {
            "_id": 579827,
            "geolocation": "-1.2625482 36.7924794 0.0 21.0",
            "meta/instanceID": "uuid:2a8129f5-3091-44e1-a579-bed2b07a12cf",
            "name": "Smith",
            "formhub/uuid": "633ec390e024411ba5ce634db7807e62",
            "_submission_time": "2013-07-03T08:25:30",
            "age": "107",
            "_uuid": "2a8129f5-3091-44e1-a579-bed2b07a12cf",
            "when": "2013-07-03",
            "amount": "250.0",
            "_geolocation": [
                "-1.2625482",
                "36.7924794"
            ],
            "_xform_id_string": "test_data_types",
            "_userform_id": "larryweya_test_data_types",
            "_status": "submitted_via_web",
            "precisely": "2013-07-03T15:24:00.000+03",
            "really": "15:24:00.000+03"
        }

        submission_2 = {
            "_id": 579828,
            "_submission_time": "2013-07-03T08:26:10",
            "_uuid": "5b4752eb-e13c-483e-87cb-e67ca6bb61e5",
            "_xform_id_string": "test_data_types",
            "_userform_id": "larryweya_test_data_types",
            "_status": "submitted_via_web",
            "meta/instanceID": "uuid:5b4752eb-e13c-483e-87cb-e67ca6bb61e5",
            "formhub/uuid": "633ec390e024411ba5ce634db7807e62",
            "amount": "",
        }

        survey = create_survey_from_xls(viewer_fixture_path(
            'test_data_types/test_data_types.xls'),
            default_name='test_data_types/test_data_types')
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        # format submission 1 for export
        survey_name = survey.name
        indices = {survey_name: 0}
        data = dict_to_joined_export(submission_1, 1, indices, survey_name)
        new_row = export_builder.pre_process_row(data[survey_name],
                                                 export_builder.sections[0])
        self.assertIsInstance(new_row['age'], int)
        self.assertIsInstance(new_row['when'], datetime.date)
        self.assertIsInstance(new_row['amount'], float)

        # check missing values dont break and empty values return blank strings
        indices = {survey_name: 0}
        data = dict_to_joined_export(submission_2, 1, indices, survey_name)
        new_row = export_builder.pre_process_row(data[survey_name],
                                                 export_builder.sections[0])
        self.assertIsInstance(new_row['amount'], str)
        self.assertEqual(new_row['amount'], '')

    def test_xls_convert_dates_before_1900(self):
        survey = create_survey_from_xls(viewer_fixture_path(
            'test_data_types/test_data_types.xls'),
            default_name='test_data_types/test_data_types.xls')
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        data = [
            {
                'name': 'Abe',
                'when': '1899-07-03',
            }
        ]
        # create export file
        temp_xls_file = NamedTemporaryFile(suffix='.xlsx')
        export_builder.to_xls_export(temp_xls_file.name, data)
        temp_xls_file.close()
        # this should error if there is a problem, not sure what to assert

    def test_convert_types(self):
        val = '1'
        expected_val = 1
        converted_val = ExportBuilder.convert_type(val, 'int')
        self.assertIsInstance(converted_val, int)
        self.assertEqual(converted_val, expected_val)

        val = '1.2'
        expected_val = 1.2
        converted_val = ExportBuilder.convert_type(val, 'decimal')
        self.assertIsInstance(converted_val, float)
        self.assertEqual(converted_val, expected_val)

        val = '2012-06-23'
        expected_val = datetime.date(2012, 6, 23)
        converted_val = ExportBuilder.convert_type(val, 'date')
        self.assertIsInstance(converted_val, datetime.date)
        self.assertEqual(converted_val, expected_val)
