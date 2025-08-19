# coding: utf-8
import csv
import datetime
import json
import os
import io
from time import sleep

import requests
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.urls import reverse
from django.utils.dateparse import parse_datetime
from openpyxl import load_workbook

from kobo.apps.openrosa.apps.main.tests.test_base import TestBase
from kobo.apps.openrosa.apps.viewer.views import (
    delete_export,
    export_list,
    create_export,
    export_progress,
    export_download,
)
from kobo.apps.openrosa.apps.viewer.xls_writer import XlsWriter
from kobo.apps.openrosa.apps.viewer.models.export import Export
from kobo.apps.openrosa.apps.viewer.models.parsed_instance import ParsedInstance
from kobo.apps.openrosa.apps.logger.models import Instance
from kobo.apps.openrosa.apps.viewer.tasks import create_xls_export
from kobo.apps.openrosa.libs.utils.export_tools import (
    generate_export,
    increment_index_in_filename,
    dict_to_joined_export,
)
from kpi.deployment_backends.kc_access.storage import (
    default_kobocat_storage as default_storage,
)

AMBULANCE_KEY = (
    'transport/available_transportation_types_to_referral_facility/ambulance'
)
AMBULANCE_KEY_DOTS = (
    'transport.available_transportation_types_to_referral_facility.ambulance'
)


def _main_fixture_path(instance_name):
    return os.path.join(settings.OPENROSA_APP_DIR, 'apps', 'main', 'tests',
                        'fixtures', 'transportation', 'instances_w_uuid',
                        instance_name, instance_name + '.xml')


class TestExports(TestBase):

    def setUp(self):
        super().setUp()
        self._submission_time = parse_datetime('2013-02-18 15:54:01Z')

    def test_unique_xls_sheet_name(self):
        xls_writer = XlsWriter()
        xls_writer.add_sheet('section9_pit_latrine_with_slab_group')
        xls_writer.add_sheet('section9_pit_latrine_without_slab_group')
        # create a set of sheet names keys
        sheet_names_set = set(xls_writer._sheets.keys())
        self.assertEqual(len(sheet_names_set), 2)

    def test_create_export(self):
        self._publish_transportation_form_and_submit_instance()
        # test xls
        export = generate_export(Export.XLS_EXPORT, 'xls', self.user.username,
                                 self.xform.id_string)
        self.assertTrue(default_storage.exists(export.filepath))
        path, ext = os.path.splitext(export.filename)
        self.assertEqual(ext, '.xls')

        # test csv
        export = generate_export(Export.CSV_EXPORT, 'csv', self.user.username,
                                 self.xform.id_string)
        self.assertTrue(default_storage.exists(export.filepath))
        path, ext = os.path.splitext(export.filename)
        self.assertEqual(ext, '.csv')

        # test xls with existing export_id
        existing_export = Export.objects.create(xform=self.xform,
                                                export_type=Export.XLS_EXPORT)
        export = generate_export(Export.XLS_EXPORT, 'xls', self.user.username,
                                 self.xform.id_string, existing_export.id)
        self.assertEqual(existing_export.id, export.id)

    def test_delete_file_on_export_delete(self):
        self._publish_transportation_form()
        self._submit_transport_instance()
        export = generate_export(
            Export.XLS_EXPORT, 'xls', self.user.username, self.xform.id_string
        )
        self.assertTrue(default_storage.exists(export.filepath))
        # delete export object
        export.delete()
        self.assertFalse(default_storage.exists(export.filepath))

    def test_graceful_exit_on_export_delete_if_file_doesnt_exist(self):
        self._publish_transportation_form()
        self._submit_transport_instance()
        export = generate_export(Export.XLS_EXPORT, 'xls', self.user.username,
                                 self.xform.id_string)

        # delete file
        default_storage.delete(export.filepath)
        self.assertFalse(default_storage.exists(export.filepath))
        # clear filename, like it would be in an incomplete export
        export.filename = None
        export.filedir = None
        export.save()
        # delete export record, which should try to delete file as well
        delete_url = reverse(delete_export, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': 'xls'
        })
        post_data = {'export_id': export.id}
        response = self.client.post(delete_url, post_data)
        self.assertEqual(response.status_code, 302)

    def test_delete_oldest_export_on_limit(self):
        self._publish_transportation_form()
        self._submit_transport_instance()
        # create first export
        first_export = generate_export(
            Export.XLS_EXPORT, 'xls', self.user.username, self.xform.id_string)
        self.assertIsNotNone(first_export.pk)
        # create exports that exceed set limit
        for i in range(Export.MAX_EXPORTS):
            generate_export(Export.XLS_EXPORT, 'xls', self.user.username,
                            self.xform.id_string)
        # first export should be deleted
        exports = Export.objects.filter(id=first_export.id)
        self.assertEqual(len(exports), 0)

    def test_create_export_url(self):
        self._publish_transportation_form()
        self._submit_transport_instance()
        num_exports = Export.objects.count()
        # create export
        create_export_url = reverse(create_export, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': Export.XLS_EXPORT
        })

        # anonymous user has to login first
        response = self.anon.post(create_export_url)
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login", response['location'])

        response = self.client.post(create_export_url)
        self.assertEqual(response.status_code, 302)
        self.assertEqual(Export.objects.count(), num_exports + 1)

    def test_delete_export_url(self):
        self._publish_transportation_form()
        self._submit_transport_instance()
        # create export
        export = generate_export(Export.XLS_EXPORT, 'xls', self.user.username,
                                 self.xform.id_string)
        exports = Export.objects.filter(id=export.id)
        self.assertEqual(len(exports), 1)
        delete_url = reverse(delete_export, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': 'xls'
        })
        post_data = {'export_id': export.id}

        # anonymous user has to login first
        response = self.anon.post(delete_url, post_data)
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login", response['location'])

        response = self.client.post(delete_url, post_data)
        self.assertEqual(response.status_code, 302)
        exports = Export.objects.filter(id=export.id)
        self.assertEqual(len(exports), 0)

    def test_export_progress_output(self):
        self._publish_transportation_form()
        self._submit_transport_instance()
        # create exports
        for i in range(2):
            generate_export(Export.XLS_EXPORT, 'xls', self.user.username,
                            self.xform.id_string)
        self.assertEqual(Export.objects.count(), 2)
        # progress for multiple exports
        progress_url = reverse(export_progress, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': 'xls'
        })
        get_data = {'export_ids': [e.id for e in Export.objects.all()]}
        response = self.client.get(progress_url, get_data)
        content = json.loads(response.content)
        self.assertEqual(len(content), 2)
        self.assertEqual(sorted(['url', 'export_id', 'complete', 'filename']),
                         sorted(content[0].keys()))

    def test_dont_auto_export_if_exports_exist(self):
        self._publish_transportation_form()
        self._submit_transport_instance()
        # create export
        create_export_url = reverse(create_export, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': Export.XLS_EXPORT
        })
        self.client.post(create_export_url)
        num_exports = Export.objects.count()
        export_list_url = reverse(export_list, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': Export.XLS_EXPORT
        })
        self.client.get(export_list_url)
        self.assertEqual(Export.objects.count(), num_exports)

    def test_last_submission_time_on_export(self):
        self._publish_transportation_form()
        self._submit_transport_instance()
        # create export
        generate_export(
            Export.XLS_EXPORT, 'xls', self.user.username, self.xform.id_string)
        num_exports = Export.objects.filter(
            xform=self.xform, export_type=Export.XLS_EXPORT).count()
        # check that our function knows there are no more submissions
        self.assertFalse(
            Export.exports_outdated(xform=self.xform,
                                    export_type=Export.XLS_EXPORT))
        sleep(1)
        # force new last submission date on xform
        last_submission = self.xform.instances.order_by('-date_created')[0]
        last_submission.date_created += datetime.timedelta(hours=1)
        last_submission.save()
        # check that our function knows data has changed
        self.assertTrue(
            Export.exports_outdated(xform=self.xform,
                                    export_type=Export.XLS_EXPORT))
        # Force a new export. Auto export has been removed in
        # https://github.com/kobotoolbox/kobocat/commit/40c67f219778065d24f405b28de790179e1fc4b2
        generate_export(
            Export.XLS_EXPORT, 'xls', self.user.username, self.xform.id_string)
        export_list_url = reverse(export_list, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': Export.XLS_EXPORT
        })
        self.client.get(export_list_url)
        self.assertEqual(
            Export.objects.filter(xform=self.xform,
                                  export_type=Export.XLS_EXPORT).count(),
            num_exports + 1)
        # Force a new export with another type. Auto export has been removed in
        # https://github.com/kobotoolbox/kobocat/commit/40c67f219778065d24f405b28de790179e1fc4b2
        num_exports = Export.objects.filter(
            xform=self.xform, export_type=Export.CSV_EXPORT).count()
        generate_export(
            Export.CSV_EXPORT, 'csv', self.user.username, self.xform.id_string)
        export_list_url = reverse(export_list, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': Export.CSV_EXPORT
        })
        self.client.get(export_list_url)
        self.assertEqual(
            Export.objects.filter(xform=self.xform,
                                  export_type=Export.CSV_EXPORT).count(),
            num_exports + 1)

    def test_last_submission_time_empty(self):
        self._publish_transportation_form()
        self._submit_transport_instance()
        # create export
        export = generate_export(Export.XLS_EXPORT, 'xls', self.user.username,
                                 self.xform.id_string)
        # set time of last submission to None
        export.time_of_last_submission = None
        export.save()
        self.assertTrue(Export.exports_outdated(xform=self.xform,
                        export_type=Export.XLS_EXPORT))

    def test_invalid_export_type(self):
        self._publish_transportation_form()
        self._submit_transport_instance()
        export_list_url = reverse(export_list, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': 'invalid'
        })
        response = self.client.get(export_list_url)
        self.assertEqual(response.status_code, 400)
        # test create url
        create_export_url = reverse(create_export, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': 'invalid'
        })
        response = self.client.post(create_export_url)
        self.assertEqual(response.status_code, 400)

    def test_add_index_to_filename(self):
        filename = "file_name-123f.txt"
        new_filename = increment_index_in_filename(filename)
        expected_filename = "file_name-123f-1.txt"
        self.assertEqual(new_filename, expected_filename)

        # test file that already has an index
        filename = "file_name-123.txt"
        new_filename = increment_index_in_filename(filename)
        expected_filename = "file_name-124.txt"
        self.assertEqual(new_filename, expected_filename)

    def test_export_download_url(self):
        self._publish_transportation_form()
        self._submit_transport_instance()
        export = generate_export(Export.CSV_EXPORT, 'csv', self.user.username,
                                 self.xform.id_string)
        csv_export_url = reverse(export_download, kwargs={
            "username": self.user.username,
            "id_string": self.xform.id_string,
            "export_type": Export.CSV_EXPORT,
            "filename": export.filename
        })
        response = self.client.get(csv_export_url)
        if not isinstance(default_storage, FileSystemStorage):
            self.assertEqual(response.status_code, 302)
        else:
            self.assertEqual(response.status_code, 200)

        # test xls
        export = generate_export(Export.XLS_EXPORT, 'xls', self.user.username,
                                 self.xform.id_string)
        xls_export_url = reverse(export_download, kwargs={
            "username": self.user.username,
            "id_string": self.xform.id_string,
            "export_type": Export.XLS_EXPORT,
            "filename": export.filename
        })
        response = self.client.get(xls_export_url)
        if not isinstance(default_storage, FileSystemStorage):
            self.assertEqual(response.status_code, 302)
        else:
            self.assertEqual(response.status_code, 200)

    def test_404_on_export_io_error(self):
        """
        Test that we return a 404 when the response_with_mimetype_and_name
        encounters an IOError
        """
        self._publish_transportation_form()
        self._submit_transport_instance()
        export = generate_export(Export.CSV_EXPORT, 'csv', self.user.username,
                                 self.xform.id_string)
        export_url = reverse(export_download, kwargs={
            "username": self.user.username,
            "id_string": self.xform.id_string,
            "export_type": Export.CSV_EXPORT,
            "filename": export.filename
        })
        # delete the export
        export.delete()
        # access the export
        response = self.client.get(export_url)
        self.assertEqual(response.status_code, 404)

    def test_deleted_submission_not_in_export(self):
        self._publish_transportation_form()
        initial_count = ParsedInstance.query_mongo(
            self.user.username, self.xform.id_string, '{}', '[]', '{}',
            count=True)[0]['count']
        self._submit_transport_instance(0)
        self._submit_transport_instance(1)
        count = ParsedInstance.query_mongo(
            self.user.username, self.xform.id_string, '{}', '[]', '{}',
            count=True)[0]['count']
        self.assertEqual(count, initial_count + 2)
        # get id of second submission
        instance_id = Instance.objects.filter(
            xform=self.xform).order_by('id').reverse()[0].id
        delete_url = reverse('data-detail', kwargs={
            'pk': self.xform.pk,
            'dataid': instance_id
        })
        self.client.delete(delete_url)
        count = ParsedInstance.query_mongo(
            self.user.username, self.xform.id_string, '{}', '[]', '{}',
            count=True)[0]['count']
        self.assertEqual(count, initial_count + 1)

        # create the export
        export = generate_export(
            Export.CSV_EXPORT, 'csv', self.user.username, self.xform.id_string
        )
        csv_export_url = reverse(export_download, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': Export.CSV_EXPORT,
            'filename': export.filename
        })
        response = self.client.get(csv_export_url)
        if not isinstance(default_storage, FileSystemStorage):
            assert response.status_code == 302
            response = requests.get(response.headers['Location'])

        assert response.status_code == 200

        f = io.StringIO(self._get_response_content(response).decode())
        csv_reader = csv.reader(f)
        num_rows = len([row for row in csv_reader])
        f.close()
        # number of rows == 2 i.e. initial_count + header plus one row
        self.assertEqual(num_rows, initial_count + 2)

    def test_edited_submissions_in_exports(self):
        self._publish_transportation_form()
        initial_count = ParsedInstance.query_mongo(
            self.user.username, self.xform.id_string, '{}', '[]', '{}',
            count=True)[0]['count']
        instance_name = 'transport_2011-07-25_19-05-36'
        path = _main_fixture_path(instance_name)
        self._make_submission(path)
        count = ParsedInstance.query_mongo(
            self.user.username, self.xform.id_string, '{}', '[]', '{}',
            count=True)[0]['count']
        self.assertEqual(count, initial_count + 1)
        # make edited submission - simulating what enketo would return
        instance_name = 'transport_2011-07-25_19-05-36-edited'
        path = _main_fixture_path(instance_name)
        self._make_submission(path)
        count = ParsedInstance.query_mongo(
            self.user.username, self.xform.id_string, '{}', '[]', '{}',
            count=True)[0]['count']
        self.assertEqual(count, initial_count + 1)
        # create the export
        export = generate_export(
            Export.CSV_EXPORT, 'csv', self.user.username, self.xform.id_string
        )
        csv_export_url = reverse(export_download, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': Export.CSV_EXPORT,
            'filename': export.filename
        })
        response = self.client.get(csv_export_url)
        if not isinstance(default_storage, FileSystemStorage):
            assert response.status_code == 302
            response = requests.get(response.headers['Location'])

        assert response.status_code == 200

        f = io.StringIO(self._get_response_content(response).decode())
        csv_reader = csv.DictReader(f)
        data = [row for row in csv_reader]
        f.close()
        num_rows = len(data)
        # number of rows == initial_count + 1
        self.assertEqual(num_rows, initial_count + 1)
        key = (
            'transport/loop_over_transport_types_frequency/ambulance/'
            'frequency_to_referral_facility'
        )
        self.assertEqual(data[initial_count][key], "monthly")

    def test_export_ids_dont_have_comma_separation(self):
        """
        It seems using {{ }} to output numbers greater than 1000 formats the
        number with a thousand separator
        """
        self._publish_transportation_form()
        self._submit_transport_instance()
        # create an in-complete export
        export = Export.objects.create(id=1234, xform=self.xform,
                                       export_type=Export.XLS_EXPORT)
        self.assertEqual(export.pk, 1234)
        export_list_url = reverse(
            export_list, kwargs={
                "username": self.user.username,
                "id_string": self.xform.id_string,
                "export_type": Export.XLS_EXPORT
            })
        response = self.client.get(export_list_url)
        self.assertContains(response, '#delete-1234"')
        self.assertNotContains(response, '#delete-1,234"')

    def test_export_progress_updates(self):
        """
        Test that after generate_export is called, we change out state to
        pending and after its complete, we change it to complete, if we fail
        between the two, updates, we have failed
        """
        self._publish_transportation_form()
        # generate an export that fails because of the NoRecordsFound exception
        export = Export.objects.create(xform=self.xform,
                                       export_type=Export.XLS_EXPORT)
        # check that progress url says pending
        progress_url = reverse(export_progress, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': 'xls'
        })
        params = {'export_ids': [export.id]}
        response = self.client.get(progress_url, params)
        status = json.loads(response.content)[0]
        self.assertEqual(status["complete"], False)
        self.assertEqual(status["filename"], None)

        export.internal_status = Export.FAILED
        export.save()
        # check that progress url says failed
        progress_url = reverse(export_progress, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': 'xls'
        })
        params = {'export_ids': [export.id]}
        response = self.client.get(progress_url, params)
        status = json.loads(response.content)[0]
        self.assertEqual(status["complete"], True)
        self.assertEqual(status["filename"], None)

        # make a submission and create a valid export
        self._submit_transport_instance()
        create_xls_export(
            self.user.username,
            self.xform.id_string, export.id)
        params = {'export_ids': [export.id]}
        response = self.client.get(progress_url, params)
        status = json.loads(response.content)[0]
        self.assertEqual(status["complete"], True)
        self.assertIsNotNone(status["filename"])

    def test_exports_outdated_doesnt_consider_failed_exports(self):
        self._publish_transportation_form()
        self._submit_transport_instance()
        # create a bad export
        export = Export.objects.create(
            xform=self.xform, export_type=Export.XLS_EXPORT,
            internal_status=Export.FAILED)
        self.assertTrue(
            Export.exports_outdated(self.xform, export.export_type))

    def test_exports_outdated_considers_pending_exports(self):
        self._publish_transportation_form()
        self._submit_transport_instance()
        # create a pending export
        export = Export.objects.create(
            xform=self.xform, export_type=Export.XLS_EXPORT,
            internal_status=Export.PENDING)
        self.assertFalse(
            Export.exports_outdated(self.xform, export.export_type))

    def _get_csv_data(self, filepath):
        csv_file = default_storage.open(filepath, mode='r')
        reader = csv.DictReader(csv_file)
        data = next(reader)
        csv_file.close()
        return data

    def _get_xls_data(self, filepath):
        with default_storage.open(filepath) as f:
            workbook = load_workbook(f)
        transportation_sheet = workbook['transportation_2011_07_25']
        self.assertTrue(transportation_sheet.max_row > 1)
        headers = [cell.value for cell in transportation_sheet[1]]
        column1 = [cell.value for cell in transportation_sheet[2]]
        return dict(zip(headers, column1))

    def test_column_header_delimiter_export_option(self):
        self._publish_transportation_form()
        # survey 1 has ambulance and bicycle as values for
        # transport/available_transportation_types_to_referral_facility
        self._submit_transport_instance(survey_at=1)
        create_csv_export_url = reverse(create_export, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': 'csv'
        })
        default_params = {}
        custom_params = {
            'options[group_delimiter]': '.',
        }
        # test csv with default group delimiter
        response = self.client.post(create_csv_export_url, default_params)
        self.assertEqual(response.status_code, 302)
        export = Export.objects.filter(
            xform=self.xform, export_type='csv').latest('created_on')
        self.assertTrue(bool(export.filepath))
        data = self._get_csv_data(export.filepath)
        self.assertTrue(AMBULANCE_KEY in data)
        self.assertEqual(data[AMBULANCE_KEY], 'True')

        sleep(1)
        # test csv with dot delimiter
        response = self.client.post(create_csv_export_url, custom_params)
        self.assertEqual(response.status_code, 302)
        export = Export.objects.filter(
            xform=self.xform, export_type='csv').latest('created_on')
        self.assertTrue(bool(export.filepath))
        data = self._get_csv_data(export.filepath)
        self.assertTrue(AMBULANCE_KEY_DOTS in data)
        self.assertEqual(data[AMBULANCE_KEY_DOTS], 'True')

        # test xls with default group delimiter
        create_csv_export_url = reverse(create_export, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': 'xls'
        })
        response = self.client.post(create_csv_export_url, default_params)
        self.assertEqual(response.status_code, 302)
        export = Export.objects.filter(
            xform=self.xform, export_type='xls').latest('created_on')
        self.assertTrue(bool(export.filepath))
        data = self._get_xls_data(export.filepath)
        self.assertTrue(AMBULANCE_KEY in data)
        self.assertTrue(data[AMBULANCE_KEY])

        sleep(1)
        # test xls with dot delimiter
        response = self.client.post(create_csv_export_url, custom_params)
        self.assertEqual(response.status_code, 302)
        export = Export.objects.filter(
            xform=self.xform, export_type='xls').latest('created_on')
        self.assertTrue(bool(export.filepath))
        data = self._get_xls_data(export.filepath)
        self.assertTrue(AMBULANCE_KEY_DOTS in data)
        self.assertTrue(data[AMBULANCE_KEY_DOTS])

    def test_split_select_multiple_export_option(self):
        self._publish_transportation_form()
        self._submit_transport_instance(survey_at=1)
        create_csv_export_url = reverse(create_export, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': 'csv'
        })
        default_params = {}
        custom_params = {
            'options[dont_split_select_multiples]': 'yes'
        }
        # test csv with default split select multiples
        response = self.client.post(create_csv_export_url, default_params)
        self.assertEqual(response.status_code, 302)
        export = Export.objects.filter(
            xform=self.xform, export_type='csv').latest('created_on')
        self.assertTrue(bool(export.filepath))
        data = self._get_csv_data(export.filepath)
        # we should have transport/available_transportation_types_to_referral_f
        # acility/ambulance as a separate column
        self.assertTrue(AMBULANCE_KEY in data)
        self.assertEqual(data[AMBULANCE_KEY], 'True')

        sleep(1)
        # test csv with default split select multiples, binary select multiples
        settings.BINARY_SELECT_MULTIPLES = True
        response = self.client.post(create_csv_export_url, default_params)
        self.assertEqual(response.status_code, 302)
        export = Export.objects.filter(
            xform=self.xform, export_type='csv').latest('created_on')
        self.assertTrue(bool(export.filepath))
        data = self._get_csv_data(export.filepath)
        # we should have transport/available_transportation_types_to_referral_f
        # acility/ambulance as a separate column
        self.assertTrue(AMBULANCE_KEY in data)
        self.assertEqual(data[AMBULANCE_KEY], '1')
        settings.BINARY_SELECT_MULTIPLES = False

        sleep(1)
        # test csv without default split select multiples
        response = self.client.post(create_csv_export_url, custom_params)
        self.assertEqual(response.status_code, 302)
        export = Export.objects.filter(
            xform=self.xform, export_type='csv').latest('created_on')
        self.assertTrue(bool(export.filepath))
        data = self._get_csv_data(export.filepath)
        # transport/available_transportation_types_to_referral_facility/ambulan
        # ce should not be in its own column
        self.assertFalse(AMBULANCE_KEY in data)
        # transport/available_transportation_types_to_referral_facility should
        # be a column
        self.assertTrue(
            'transport/available_transportation_types_to_referral_facility' in
            data)
        # check that ambulance is one the values within the transport/available
        # _transportation_types_to_referral_facility column
        self.assertTrue("ambulance" in data[
            'transport/available_transportation_types_to_referral_facility'
        ].split(" "))

        create_xls_export_url = reverse(create_export, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': 'xls'
        })
        # test xls with default split select multiples
        response = self.client.post(create_xls_export_url, default_params)
        self.assertEqual(response.status_code, 302)
        export = Export.objects.filter(
            xform=self.xform, export_type='xls').latest('created_on')
        self.assertTrue(bool(export.filepath))
        data = self._get_xls_data(export.filepath)
        # we should have transport/available_transportation_types_to_referral_f
        # acility/ambulance as a separate column
        self.assertTrue(AMBULANCE_KEY in data)

        sleep(1)
        # test xls without default split select multiples
        response = self.client.post(create_xls_export_url, custom_params)
        self.assertEqual(response.status_code, 302)
        export = Export.objects.filter(
            xform=self.xform, export_type='xls').latest('created_on')
        self.assertTrue(bool(export.filepath))
        data = self._get_xls_data(export.filepath)
        # transport/available_transportation_types_to_referral_facility/ambulan
        # ce should NOT be in its own column
        self.assertFalse(AMBULANCE_KEY in data)
        # transport/available_transportation_types_to_referral_facility should
        # be a column
        self.assertTrue(
            'transport/available_transportation_types_to_referral_facility'
            in data)
        # check that ambulance is one the values within the transport/available
        # _transportation_types_to_referral_facility column
        self.assertTrue("ambulance" in data[
            'transport/available_transportation_types_to_referral_facility'
        ].split(" "))

    def test_dict_to_joined_export_works(self):
        data =\
            {
                'name': 'Abe',
                'age': '35',
                '_geolocation': [None, None],
                'attachments': ['abcd.jpg', 'efgh.jpg'],
                'children':
                [
                    {
                        'children/name': 'Mike',
                        'children/age': '5',
                        'children/cartoons':
                        [
                            {
                                'children/cartoons/name': 'Tom & Jerry',
                                'children/cartoons/why': 'Tom is silly',
                            },
                            {
                                'children/cartoons/name': 'Flinstones',
                                'children/cartoons/why':
                                "I like bamb bam\u0107",
                            }
                        ]
                    },
                    {
                        'children/name': 'John',
                        'children/age': '2',
                        'children/cartoons': []
                    },
                    {
                        'children/name': 'Imora',
                        'children/age': '3',
                        'children/cartoons':
                        [
                            {
                                'children/cartoons/name': 'Shrek',
                                'children/cartoons/why': 'He\'s so funny'
                            },
                            {
                                'children/cartoons/name': 'Dexter\'s Lab',
                                'children/cartoons/why': 'He thinks hes smart',
                                'children/cartoons/characters':
                                [
                                    {
                                        'children/cartoons/characters/name':
                                        'Dee Dee',
                                        'children/cartoons/characters/good_or_'
                                        'evil': 'good'
                                    },
                                    {
                                        'children/cartoons/characters/name':
                                        'Dexter',
                                        'children/cartoons/characters/good_or_'
                                        'evil': 'evil'
                                    },
                                ]
                            }
                        ]
                    }
                ]
            }
        expected_output =\
            {
                'survey': {
                    'name': 'Abe',
                    'age': '35'
                },
                'children':
                [
                    {
                        'children/name': 'Mike',
                        'children/age': '5',
                        '_index': 1,
                        '_parent_table_name': 'survey',
                        '_parent_index': 1
                    },
                    {
                        'children/name': 'John',
                        'children/age': '2',
                        '_index': 2,
                        '_parent_table_name': 'survey',
                        '_parent_index': 1
                    },
                    {
                        'children/name': 'Imora',
                        'children/age': '3',
                        '_index': 3,
                        '_parent_table_name': 'survey',
                        '_parent_index': 1
                    },
                ],
                'children/cartoons':
                [
                    {
                        'children/cartoons/name': 'Tom & Jerry',
                        'children/cartoons/why': 'Tom is silly',
                        '_index': 1,
                        '_parent_table_name': 'children',
                        '_parent_index': 1
                    },
                    {
                        'children/cartoons/name': 'Flinstones',
                        'children/cartoons/why': "I like bamb bam\u0107",
                        '_index': 2,
                        '_parent_table_name': 'children',
                        '_parent_index': 1
                    },
                    {
                        'children/cartoons/name': 'Shrek',
                        'children/cartoons/why': 'He\'s so funny',
                        '_index': 3,
                        '_parent_table_name': 'children',
                        '_parent_index': 3
                    },
                    {
                        'children/cartoons/name': 'Dexter\'s Lab',
                        'children/cartoons/why': 'He thinks hes smart',
                        '_index': 4,
                        '_parent_table_name': 'children',
                        '_parent_index': 3
                    }
                ],
                'children/cartoons/characters':
                [
                    {
                        'children/cartoons/characters/name': 'Dee Dee',
                        'children/cartoons/characters/good_or_evil': 'good',
                        '_index': 1,
                        '_parent_table_name': 'children/cartoons',
                        '_parent_index': 4
                    },
                    {
                        'children/cartoons/characters/name': 'Dexter',
                        'children/cartoons/characters/good_or_evil': 'evil',
                        '_index': 2,
                        '_parent_table_name': 'children/cartoons',
                        '_parent_index': 4
                    }
                ]
            }
        survey_name = 'survey'
        indices = {survey_name: 0}
        output = dict_to_joined_export(data, 1, indices, survey_name)
        self.assertEqual(output[survey_name], expected_output[survey_name])
        # 1st level
        self.assertEqual(len(output['children']), 3)
        for child in enumerate(['Mike', 'John', 'Imora']):
            index = child[0]
            name = child[1]
            self.assertEqual(
                [x for x in output['children']
                 if x['children/name'] == name][0],
                expected_output['children'][index])
        # 2nd level
        self.assertEqual(len(output['children/cartoons']), 4)
        for cartoon in enumerate(
                ['Tom & Jerry', 'Flinstones', 'Shrek', 'Dexter\'s Lab']):
            index = cartoon[0]
            name = cartoon[1]
            self.assertEqual(
                [x for x in output['children/cartoons']
                 if x['children/cartoons/name'] == name][0],
                expected_output['children/cartoons'][index])
        # 3rd level
        self.assertEqual(len(output['children/cartoons/characters']), 2)
        for characters in enumerate(['Dee Dee', 'Dexter']):
            index = characters[0]
            name = characters[1]
            self.assertEqual(
                [x for x in output['children/cartoons/characters']
                 if x['children/cartoons/characters/name'] == name][0],
                expected_output['children/cartoons/characters'][index])

    def test_dict_to_joined_export_notes(self):
        submission = {
            "_id": 579828,
            "_submission_time": "2013-07-03T08:26:10",
            "_uuid": "5b4752eb-e13c-483e-87cb-e67ca6bb61e5",
            "_xform_id_string": "test_data_types",
            "_userform_id": "larryweya_test_data_types",
            "_status": "submitted_via_web",
            "_notes": [
                {
                    "note": "Note 1",
                    "date_created": "2013-07-03T08:26:10",
                    "id": 356,
                    "date_modified": "2013-07-03T08:26:10"
                },
                {
                    "note": "Note 2",
                    "date_created": "2013-07-03T08:34:40",
                    "id": 357,
                    "date_modified": "2013-07-03T08:34:40"
                },
                {
                    "note": "Note 3",
                    "date_created": "2013-07-03T08:56:14",
                    "id": 361,
                    "date_modified": "2013-07-03T08:56:14"
                }
            ],
            "meta/instanceID": "uuid:5b4752eb-e13c-483e-87cb-e67ca6bb61e5",
            "formhub/uuid": "633ec390e024411ba5ce634db7807e62",
            "amount": "",
        }

        survey_name = 'tutorial'
        indices = {survey_name: 0}
        data = dict_to_joined_export(submission, 1, indices, survey_name)
        expected_data = {
            'tutorial': {
                '_id': 579828,
                '_submission_time': '2013-07-03T08:26:10',
                '_uuid': '5b4752eb-e13c-483e-87cb-e67ca6bb61e5',
                'amount': '',
                '_xform_id_string': 'test_data_types',
                '_userform_id': 'larryweya_test_data_types',
                '_status': 'submitted_via_web',
                '_notes': 'Note 1\nNote 2\nNote 3',
                'meta/instanceID': 'uuid:5b4752eb-e13c-483e-87cb-e67ca6bb61e5',
                'formhub/uuid': '633ec390e024411ba5ce634db7807e62'
            }
        }
        self.assertEqual(sorted(data), sorted(expected_data))

    def test_create_xls_export_non_existent_id(self):
        self._publish_transportation_form()

        # make a submission and create a valid export
        self._submit_transport_instance()
        non_existent_id = 42
        result = create_xls_export(
            self.user.username,
            self.xform.id_string, non_existent_id)

        self.assertEqual(result, None)
