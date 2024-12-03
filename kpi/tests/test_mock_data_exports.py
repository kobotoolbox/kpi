# flake8: noqa: F401
import datetime
import os
import zipfile
from collections import defaultdict
from unittest import mock
from zoneinfo import ZoneInfo

import openpyxl
from django.conf import settings
from django.test import TestCase
from django.urls import reverse

from kobo.apps.kobo_auth.shortcuts import User
from kobo.apps.reports import report_data
from kpi.constants import (
    PERM_CHANGE_ASSET,
    PERM_PARTIAL_SUBMISSIONS,
    PERM_VIEW_ASSET,
    PERM_VIEW_SUBMISSIONS,
)
from kpi.models import Asset, SubmissionExportTask
from kpi.utils.mongo_helper import drop_mock_only
from kpi.utils.object_permission import get_anonymous_user


class MockDataExportsBase(TestCase):
    """
    Creates self.asset, deploys it using the mock backend, and makes some
    submissions to it
    """

    fixtures = ['test_data']

    forms = {
        'Identificación de animales': {
            'content': {
                'choices': [
                    {
                        '$autovalue': 'spherical',
                        '$kuid': 'jfDnpH2n9',
                        'label': ['Spherical', 'Esf\xe9rico'],
                        'list_name': 'symmetry',
                        'name': 'spherical',
                    },
                    {
                        '$autovalue': 'radial',
                        '$kuid': '07Wr5ehxt',
                        'label': ['Radial', 'Radial'],
                        'list_name': 'symmetry',
                        'name': 'radial',
                    },
                    {
                        '$autovalue': 'bilateral',
                        '$kuid': 'vn5m3TZkF',
                        'label': ['Bilateral', 'Bilateral'],
                        'list_name': 'symmetry',
                        'name': 'bilateral',
                    },
                    {
                        '$autovalue': 'yes',
                        '$kuid': 'thb0B532I',
                        'label': ['Yes', 'S\xed'],
                        'list_name': 'fluids',
                        'name': 'yes',
                    },
                    {
                        '$autovalue': 'yes__and_some_',
                        '$kuid': 'WrTMETvzY',
                        'label': [
                            'Yes, and some extracellular space',
                            'S\xed, y alg\xfan espacio extracelular',
                        ],
                        'list_name': 'fluids',
                        'name': 'yes__and_some_',
                    },
                    {
                        '$autovalue': 'no___unsure',
                        '$kuid': 'i1KtAuy3a',
                        'label': ['No / Unsure', 'No / Inseguro'],
                        'list_name': 'fluids',
                        'name': 'no___unsure',
                    },
                    {
                        '$autovalue': 'yes',
                        '$kuid': 'QfrYJgSNH',
                        'label': ['Yes', 'S\xed'],
                        'list_name': 'yes_no',
                        'name': 'yes',
                    },
                    {
                        '$autovalue': 'no',
                        '$kuid': 'KzgCswpU2',
                        'label': ['No', 'No'],
                        'list_name': 'yes_no',
                        'name': 'no',
                    },
                ],
                'schema': '1',
                'settings': {'id_string': 'Identificaci_n_de_animales'},
                'survey': [
                    {
                        '$autoname': 'start',
                        '$kuid': 'df516ecd',
                        'name': 'start',
                        'type': 'start',
                    },
                    {
                        '$autoname': 'end',
                        '$kuid': '7b054499',
                        'name': 'end',
                        'type': 'end',
                    },
                    {
                        '$autoname': 'external_characteristics',
                        '$kuid': 'cbc4ba77',
                        'label': [
                            'External Characteristics',
                            'Caracter\xedsticas externas',
                        ],
                        'name': 'external_characteristics',
                        'type': 'begin_group',
                    },
                    {
                        '$autoname': 'What_kind_of_symmetry_do_you_have',
                        '$kuid': 'f073bdb4',
                        'label': [
                            'What kind of symmetry do you have?',
                            '\xbfQu\xe9 tipo de simetr\xeda tiene?',
                        ],
                        'name': 'What_kind_of_symmetry_do_you_have',
                        'required': False,
                        'select_from_list_name': 'symmetry',
                        'tags': ['hxl:#symmetry'],
                        'type': 'select_multiple',
                    },
                    {
                        '$autoname': 'How_many_segments_does_your_body_have',
                        '$kuid': '2b4d8728',
                        'label': [
                            'How many segments does your body have?',
                            '\xbfCu\xe1ntos segmentos tiene tu cuerpo?',
                        ],
                        'name': 'How_many_segments_does_your_body_have',
                        'required': False,
                        'tags': ['hxl:#segments'],
                        'type': 'integer',
                    },
                    {'$kuid': '56d0cd68', 'type': 'end_group'},
                    {
                        '$autoname': 'Do_you_have_body_flu_intracellular_space',
                        '$kuid': '5fa1fc59',
                        'label': [
                            'Do you have body fluids that occupy intracellular space?',
                            '\xbfTienes fluidos corporales que ocupan espacio intracelular?',
                        ],
                        'name': 'Do_you_have_body_flu_intracellular_space',
                        'required': False,
                        'select_from_list_name': 'fluids',
                        'tags': ['hxl:#fluids'],
                        'type': 'select_one',
                    },
                    {
                        '$autoname': 'Do_you_descend_from_unicellular_organism',
                        '$kuid': 'bfde6907',
                        'label': [
                            'Do you descend from an ancestral unicellular organism?',
                            '\xbfDesciende de un organismo unicelular ancestral?',
                        ],
                        'name': 'Do_you_descend_from_unicellular_organism',
                        'required': False,
                        'select_from_list_name': 'yes_no',
                        'type': 'select_one',
                    },
                ],
                'translated': ['label'],
                'translations': ['English', 'Spanish'],
            },
            'submissions': [
                {
                    'Do_you_descend_from_unicellular_organism': 'no',
                    'Do_you_have_body_flu_intracellular_space': 'yes__and_some_',
                    '_attachments': [],
                    '_bamboo_dataset_id': '',
                    '_geolocation': [None, None],
                    '_notes': [],
                    '_status': 'submitted_via_web',
                    '_submission_time': '2017-10-23T09:41:19',
                    '_submitted_by': None,
                    '_tags': [],
                    '_uuid': '48583952-1892-4931-8d9c-869e7b49bafb',
                    '_xform_id_string': 'aX6CUrtnHfZE64CnNdjzuz',
                    'end': '2017-10-23T05:41:13.000-04:00',
                    'external_characteristics/How_many_segments_does_your_body_have': '6',
                    'external_characteristics/What_kind_of_symmetry_do_you_have': 'spherical radial bilateral',
                    'formhub/uuid': '1511083383a64c9dad1eca3795cd3788',
                    'meta/instanceID': 'uuid:48583952-1892-4931-8d9c-869e7b49bafb',
                    'start': '2017-10-23T05:40:39.000-04:00',
                },
                {
                    'Do_you_descend_from_unicellular_organism': 'no',
                    'Do_you_have_body_flu_intracellular_space': 'yes',
                    '_attachments': [],
                    '_bamboo_dataset_id': '',
                    '_geolocation': [None, None],
                    '_notes': [],
                    '_status': 'submitted_via_web',
                    '_submission_time': '2017-10-23T09:41:38',
                    '_submitted_by': None,
                    '_tags': [],
                    '_uuid': '317ba7b7-bea4-4a8c-8620-a483c3079c4b',
                    '_xform_id_string': 'aX6CUrtnHfZE64CnNdjzuz',
                    'end': '2017-10-23T05:41:32.000-04:00',
                    'external_characteristics/How_many_segments_does_your_body_have': '3',
                    'external_characteristics/What_kind_of_symmetry_do_you_have': 'radial',
                    'formhub/uuid': '1511083383a64c9dad1eca3795cd3788',
                    'meta/instanceID': 'uuid:317ba7b7-bea4-4a8c-8620-a483c3079c4b',
                    'start': '2017-10-23T05:41:14.000-04:00',
                },
                {
                    'Do_you_descend_from_unicellular_organism': 'yes',
                    'Do_you_have_body_flu_intracellular_space': 'no___unsure',
                    '_attachments': [],
                    '_bamboo_dataset_id': '',
                    '_geolocation': [None, None],
                    '_notes': [],
                    '_status': 'submitted_via_web',
                    '_submission_time': '2017-10-23T09:42:11',
                    '_submitted_by': 'anotheruser',
                    '_tags': [],
                    '_uuid': '3f15cdfe-3eab-4678-8352-7806febf158d',
                    '_xform_id_string': 'aX6CUrtnHfZE64CnNdjzuz',
                    'end': '2017-10-23T05:42:05.000-04:00',
                    'external_characteristics/How_many_segments_does_your_body_have': '2',
                    'external_characteristics/What_kind_of_symmetry_do_you_have': 'bilateral',
                    'formhub/uuid': '1511083383a64c9dad1eca3795cd3788',
                    'meta/instanceID': 'uuid:3f15cdfe-3eab-4678-8352-7806febf158d',
                    'start': '2017-10-23T05:41:32.000-04:00',
                },
            ],
        },
        'Simple repeat group': {
            'content': {
                'schema': '1',
                'survey': [
                    {
                        'type': 'begin_group',
                        'name': 'people',
                        'label': ['People'],
                    },
                    {
                        'type': 'begin_repeat',
                        'name': 'person',
                        'label': ['person'],
                    },
                    {
                        'type': 'text',
                        'label': ['name'],
                    },
                    {
                        'type': 'integer',
                        'label': ['age'],
                    },
                    {'type': 'end_repeat'},
                    {'type': 'end_group'},
                ],
                'settings': {},
                'translated': ['label'],
                'translations': [None],
            },
            'submissions': [
                {
                    'formhub/uuid': 'cfb562511e8e44d1998de69002b492d9',
                    'people/person': [
                        {
                            'people/person/name': 'Julius Caesar',
                            'people/person/age': '55',
                        },
                        {
                            'people/person/name': 'Augustus',
                            'people/person/age': '75',
                        },
                    ],
                    '__version__': 'vbKavWWCpgBCZms6hQX4FN',
                    'meta/instanceID': 'uuid:f80be949-89b5-4af1-a29d-7d292b2bc0cd',
                    '_xform_id_string': 'aaURCfR8mYe8pzc5h3YiZG',
                    '_uuid': 'f80be949-89b5-4af1-a29d-7d292b2bc0cd',
                    '_attachments': [],
                    '_status': 'submitted_via_web',
                    '_geolocation': [None, None],
                    '_submission_time': '2021-06-30T22:12:56',
                    '_tags': [],
                    '_notes': [],
                    '_validation_status': {},
                    '_submitted_by': None,
                }
            ],
        },
        'Simple media': {
            'content': {
                'schema': '1',
                'survey': [
                    {
                        'type': 'image',
                        'name': 'an_image',
                        'label': ['Submit an image'],
                    },
                ],
                'settings': {},
                'translated': ['label'],
                'translations': [None],
            },
            'submissions': [
                {
                    'formhub/uuid': 'cfb562511e8e44d1998de69002b49299',
                    'an_image': 'audio_conversion_test_image.jpg',
                    '__version__': 'vbKavWWCpgBCZms6hQX4FB',
                    'meta/instanceID': 'uuid:f80be949-89b5-4af1-a42d-7d292b2bc0cd',
                    '_xform_id_string': 'aaURCfR8mYe8pzc5h3YiZz',
                    '_uuid': 'f80be949-89b5-4af1-a42d-7d292b2bc0cd',
                    '_attachments': [
                        {
                            'download_url': 'http://testserver/audio_conversion_test_image.jpg',
                            'filename': 'path/to/audio_conversion_test_image.jpg',
                        }
                    ],
                    '_status': 'submitted_via_web',
                    '_geolocation': [None, None],
                    '_submission_time': '2021-06-30T22:12:56',
                    '_tags': [],
                    '_notes': [],
                    '_validation_status': {},
                    '_submitted_by': None,
                }
            ],
        },
    }

    @drop_mock_only
    def setUp(self):
        self.user = User.objects.get(username='someuser')
        self.form_names = list(self.forms.keys())
        # Clean up MongoDB documents
        settings.MONGO_DB.instances.drop()

        self.assets = {
            name: self._create_asset_with_submissions(
                user=self.user,
                content=form['content'],
                name=name,
                submissions=form['submissions'],
            )
            for name, form in self.forms.items()
        }
        self.asset = self.assets[self.form_names[0]]

    @staticmethod
    def _create_asset_with_submissions(user, content, name, submissions):
        asset = Asset.objects.create(
            name=name,
            content=content,
            owner=user
        )
        asset.deploy(backend='mock', active=True)
        asset.save()

        v_uid = asset.latest_deployed_version.uid
        for submission in submissions:
            submission.update({
                '__version__': v_uid
            })
        asset.deployment.set_namespace('api_v2')
        asset.deployment.mock_submissions(submissions)
        return asset


class MockDataExports(MockDataExportsBase):
    def setUp(self):
        super().setUp()

        self.anotheruser = User.objects.get(username='anotheruser')
        partial_perms = {
            PERM_VIEW_SUBMISSIONS: [
                {'_submitted_by': self.anotheruser.username}
            ]
        }
        for asset in self.assets.values():
            asset.assign_perm(
                self.anotheruser,
                PERM_PARTIAL_SUBMISSIONS,
                partial_perms=partial_perms,
            )

        self.formpack, self.submission_stream = report_data.build_formpack(
            self.asset,
            submission_stream=self.asset.deployment.get_submissions(
                self.asset.owner
            ),
        )

    def run_csv_export_test(
        self, expected_lines=None, export_options=None, asset=None, user=None
    ):
        """
        Repeat yourself less while writing CSV export tests.

        `expected_lines`: a list of strings *without* trailing newlines whose
                          UTF-8 encoded representation should match the export
                          result
        `export_options`: (optional) a list of extra options for
                          `SubmissionExportTask.data`. Do not include `source` or `type`
        `asset`: (optional) the asset to export. Defaults to `self.asset`
        `user`: (optional) the user to own the export. Defaults to `self.user`
        """
        export_task = SubmissionExportTask()
        asset = self.asset if asset is None else asset
        export_task.user = self.user if user is None else user
        export_task.data = {
            'source': reverse('asset-detail', args=[asset.uid]),
            'type': 'csv'
        }
        if export_options:
            export_task.data.update(export_options)
        messages = defaultdict(list)
        export_task._run_task(messages)

        if expected_lines is not None:
            expected_lines = [
                (line + '\r\n').encode('utf-8') for line in expected_lines
            ]
            result_lines = list(export_task.result)
            self.assertEqual(result_lines, expected_lines)

        self.assertFalse(messages)

    def run_xls_export_test(
        self,
        expected_data,
        export_options=None,
        user=None,
        asset=None,
        repeat_group=False,
    ):
        """
        Repeat yourself less while writing XLS export tests.
        `expected_rows`: a list of strings *without* trailing newlines whose
                          UTF-8 encoded representation should match the export
                          result
        `export_options`: a list of extra options for `SubmissionExportTask.data`. Do not
                          include `source` or `type`
        """
        asset = self.asset if asset is None else asset
        export_task = SubmissionExportTask()
        export_task.user = self.user if user is None else user
        export_task.data = {
            'source': reverse('asset-detail', args=[asset.uid]),
            'type': 'xls',
        }
        if export_options:
            export_task.data.update(export_options)
        messages = defaultdict(list)
        export_task._run_task(messages)
        assert not messages

        book = openpyxl.load_workbook(export_task.result)
        expected_sheet_names = list(expected_data.keys())
        assert book.sheetnames == expected_sheet_names

        for sheet_name in expected_sheet_names:
            expected_rows = expected_data[sheet_name]
            sheet = book[sheet_name]
            assert sheet.max_row == len(expected_rows)

            for row_index, expected_row in enumerate(expected_rows):
                result_row = [
                    cell.value if cell.value is not None else ''
                    for cell in sheet[row_index + 1]
                ]
                assert result_row == expected_row

    def test_csv_export_default_options(self):
        submissions = self.forms[self.form_names[0]]['submissions']
        version_uid = self.asset.latest_deployed_version.uid
        expected_lines = [
            '"start";"end";"What kind of symmetry do you have?";"What kind of symmetry do you have?/Spherical";"What kind of symmetry do you have?/Radial";"What kind of symmetry do you have?/Bilateral";"How many segments does your body have?";"Do you have body fluids that occupy intracellular space?";"Do you descend from an ancestral unicellular organism?";"_id";"_uuid";"_submission_time";"_validation_status";"_notes";"_status";"_submitted_by";"__version__";"_tags";"_index"',
            '"";"";"#symmetry";"";"";"";"#segments";"#fluids";"";"";"";"";"";"";"";"";"";"";""',
            f'"2017-10-23T05:40:39.000-04:00";"2017-10-23T05:41:13.000-04:00";"Spherical Radial Bilateral";"1";"1";"1";"6";"Yes, and some extracellular space";"No";"{submissions[0]["_id"]}";"48583952-1892-4931-8d9c-869e7b49bafb";"2017-10-23T09:41:19";"";"";"submitted_via_web";"";"{version_uid}";"";"1"',
            f'"2017-10-23T05:41:14.000-04:00";"2017-10-23T05:41:32.000-04:00";"Radial";"0";"1";"0";"3";"Yes";"No";"{submissions[1]["_id"]}";"317ba7b7-bea4-4a8c-8620-a483c3079c4b";"2017-10-23T09:41:38";"";"";"submitted_via_web";"";"{version_uid}";"";"2"',
            f'"2017-10-23T05:41:32.000-04:00";"2017-10-23T05:42:05.000-04:00";"Bilateral";"0";"0";"1";"2";"No / Unsure";"Yes";"{submissions[2]["_id"]}";"3f15cdfe-3eab-4678-8352-7806febf158d";"2017-10-23T09:42:11";"";"";"submitted_via_web";"anotheruser";"{version_uid}";"";"3"',
        ]
        self.run_csv_export_test(expected_lines)

    def test_csv_export_default_options_partial_submissions(self):
        submissions = self.forms[self.form_names[0]]['submissions']
        version_uid = self.asset.latest_deployed_version_uid
        expected_lines = [
            '"start";"end";"What kind of symmetry do you have?";"What kind of symmetry do you have?/Spherical";"What kind of symmetry do you have?/Radial";"What kind of symmetry do you have?/Bilateral";"How many segments does your body have?";"Do you have body fluids that occupy intracellular space?";"Do you descend from an ancestral unicellular organism?";"_id";"_uuid";"_submission_time";"_validation_status";"_notes";"_status";"_submitted_by";"__version__";"_tags";"_index"',
            f'"";"";"#symmetry";"";"";"";"#segments";"#fluids";"";"";"";"";"";"";"";"";"";"";""',
            f'"2017-10-23T05:41:32.000-04:00";"2017-10-23T05:42:05.000-04:00";"Bilateral";"0";"0";"1";"2";"No / Unsure";"Yes";"{submissions[2]["_id"]}";"3f15cdfe-3eab-4678-8352-7806febf158d";"2017-10-23T09:42:11";"";"";"submitted_via_web";"anotheruser";"{version_uid}";"";"1"',
        ]
        self.run_csv_export_test(expected_lines, user=self.anotheruser)

    def test_csv_export_english_labels(self):
        submissions = self.forms[self.form_names[0]]['submissions']
        version_uid = self.asset.latest_deployed_version_uid
        export_options = {
            'lang': 'English',
        }
        expected_lines = [
            '"start";"end";"What kind of symmetry do you have?";"What kind of symmetry do you have?/Spherical";"What kind of symmetry do you have?/Radial";"What kind of symmetry do you have?/Bilateral";"How many segments does your body have?";"Do you have body fluids that occupy intracellular space?";"Do you descend from an ancestral unicellular organism?";"_id";"_uuid";"_submission_time";"_validation_status";"_notes";"_status";"_submitted_by";"__version__";"_tags";"_index"',
            f'"";"";"#symmetry";"";"";"";"#segments";"#fluids";"";"";"";"";"";"";"";"";"";"";""',
            f'"2017-10-23T05:40:39.000-04:00";"2017-10-23T05:41:13.000-04:00";"Spherical Radial Bilateral";"1";"1";"1";"6";"Yes, and some extracellular space";"No";"{submissions[0]["_id"]}";"48583952-1892-4931-8d9c-869e7b49bafb";"2017-10-23T09:41:19";"";"";"submitted_via_web";"";"{version_uid}";"";"1"',
            f'"2017-10-23T05:41:14.000-04:00";"2017-10-23T05:41:32.000-04:00";"Radial";"0";"1";"0";"3";"Yes";"No";"{submissions[1]["_id"]}";"317ba7b7-bea4-4a8c-8620-a483c3079c4b";"2017-10-23T09:41:38";"";"";"submitted_via_web";"";"{version_uid}";"";"2"',
            f'"2017-10-23T05:41:32.000-04:00";"2017-10-23T05:42:05.000-04:00";"Bilateral";"0";"0";"1";"2";"No / Unsure";"Yes";"{submissions[2]["_id"]}";"3f15cdfe-3eab-4678-8352-7806febf158d";"2017-10-23T09:42:11";"";"";"submitted_via_web";"anotheruser";"{version_uid}";"";"3"',
        ]
        self.run_csv_export_test(expected_lines, export_options)

    def test_csv_export_spanish_labels(self):
        submissions = self.forms[self.form_names[0]]['submissions']
        version_uid = self.asset.latest_deployed_version_uid
        export_options = {
            'lang': 'Spanish',
        }
        expected_lines = [
            '"start";"end";"¿Qué tipo de simetría tiene?";"¿Qué tipo de simetría tiene?/Esférico";"¿Qué tipo de simetría tiene?/Radial";"¿Qué tipo de simetría tiene?/Bilateral";"¿Cuántos segmentos tiene tu cuerpo?";"¿Tienes fluidos corporales que ocupan espacio intracelular?";"¿Desciende de un organismo unicelular ancestral?";"_id";"_uuid";"_submission_time";"_validation_status";"_notes";"_status";"_submitted_by";"__version__";"_tags";"_index"',
            '"";"";"#symmetry";"";"";"";"#segments";"#fluids";"";"";"";"";"";"";"";"";"";"";""',
            f'"2017-10-23T05:40:39.000-04:00";"2017-10-23T05:41:13.000-04:00";"Esférico Radial Bilateral";"1";"1";"1";"6";"Sí, y algún espacio extracelular";"No";"{submissions[0]["_id"]}";"48583952-1892-4931-8d9c-869e7b49bafb";"2017-10-23T09:41:19";"";"";"submitted_via_web";"";"{version_uid}";"";"1"',
            f'"2017-10-23T05:41:14.000-04:00";"2017-10-23T05:41:32.000-04:00";"Radial";"0";"1";"0";"3";"Sí";"No";"{submissions[1]["_id"]}";"317ba7b7-bea4-4a8c-8620-a483c3079c4b";"2017-10-23T09:41:38";"";"";"submitted_via_web";"";"{version_uid}";"";"2"',
            f'"2017-10-23T05:41:32.000-04:00";"2017-10-23T05:42:05.000-04:00";"Bilateral";"0";"0";"1";"2";"No / Inseguro";"Sí";"{submissions[2]["_id"]}";"3f15cdfe-3eab-4678-8352-7806febf158d";"2017-10-23T09:42:11";"";"";"submitted_via_web";"anotheruser";"{version_uid}";"";"3"',
        ]
        self.run_csv_export_test(expected_lines, export_options)

    def test_csv_export_english_labels_no_hxl(self):
        submissions = self.forms[self.form_names[0]]['submissions']
        version_uid = self.asset.latest_deployed_version_uid
        export_options = {
            'lang': 'English',
            'tag_cols_for_header': [],
        }
        expected_lines = [
            '"start";"end";"What kind of symmetry do you have?";"What kind of symmetry do you have?/Spherical";"What kind of symmetry do you have?/Radial";"What kind of symmetry do you have?/Bilateral";"How many segments does your body have?";"Do you have body fluids that occupy intracellular space?";"Do you descend from an ancestral unicellular organism?";"_id";"_uuid";"_submission_time";"_validation_status";"_notes";"_status";"_submitted_by";"__version__";"_tags";"_index"',
            f'"2017-10-23T05:40:39.000-04:00";"2017-10-23T05:41:13.000-04:00";"Spherical Radial Bilateral";"1";"1";"1";"6";"Yes, and some extracellular space";"No";"{submissions[0]["_id"]}";"48583952-1892-4931-8d9c-869e7b49bafb";"2017-10-23T09:41:19";"";"";"submitted_via_web";"";"{version_uid}";"";"1"',
            f'"2017-10-23T05:41:14.000-04:00";"2017-10-23T05:41:32.000-04:00";"Radial";"0";"1";"0";"3";"Yes";"No";"{submissions[1]["_id"]}";"317ba7b7-bea4-4a8c-8620-a483c3079c4b";"2017-10-23T09:41:38";"";"";"submitted_via_web";"";"{version_uid}";"";"2"',
            f'"2017-10-23T05:41:32.000-04:00";"2017-10-23T05:42:05.000-04:00";"Bilateral";"0";"0";"1";"2";"No / Unsure";"Yes";"{submissions[2]["_id"]}";"3f15cdfe-3eab-4678-8352-7806febf158d";"2017-10-23T09:42:11";"";"";"submitted_via_web";"anotheruser";"{version_uid}";"";"3"',
        ]
        self.run_csv_export_test(expected_lines, export_options)

    def test_csv_export_english_labels_group_sep(self):
        submissions = self.forms[self.form_names[0]]['submissions']
        version_uid = self.asset.latest_deployed_version_uid
        # Check `group_sep` by looking at the `select_multiple` question
        export_options = {
            'lang': 'English',
            'group_sep': '%',
        }
        expected_lines = [
            '"start";"end";"What kind of symmetry do you have?";"What kind of symmetry do you have?%Spherical";"What kind of symmetry do you have?%Radial";"What kind of symmetry do you have?%Bilateral";"How many segments does your body have?";"Do you have body fluids that occupy intracellular space?";"Do you descend from an ancestral unicellular organism?";"_id";"_uuid";"_submission_time";"_validation_status";"_notes";"_status";"_submitted_by";"__version__";"_tags";"_index"',
            '"";"";"#symmetry";"";"";"";"#segments";"#fluids";"";"";"";"";"";"";"";"";"";"";""',
            f'"2017-10-23T05:40:39.000-04:00";"2017-10-23T05:41:13.000-04:00";"Spherical Radial Bilateral";"1";"1";"1";"6";"Yes, and some extracellular space";"No";"{submissions[0]["_id"]}";"48583952-1892-4931-8d9c-869e7b49bafb";"2017-10-23T09:41:19";"";"";"submitted_via_web";"";"{version_uid}";"";"1"',
            f'"2017-10-23T05:41:14.000-04:00";"2017-10-23T05:41:32.000-04:00";"Radial";"0";"1";"0";"3";"Yes";"No";"{submissions[1]["_id"]}";"317ba7b7-bea4-4a8c-8620-a483c3079c4b";"2017-10-23T09:41:38";"";"";"submitted_via_web";"";"{version_uid}";"";"2"',
            f'"2017-10-23T05:41:32.000-04:00";"2017-10-23T05:42:05.000-04:00";"Bilateral";"0";"0";"1";"2";"No / Unsure";"Yes";"{submissions[2]["_id"]}";"3f15cdfe-3eab-4678-8352-7806febf158d";"2017-10-23T09:42:11";"";"";"submitted_via_web";"anotheruser";"{version_uid}";"";"3"',
        ]
        self.run_csv_export_test(expected_lines, export_options)

    def test_csv_export_hierarchy_in_labels(self):
        submissions = self.forms[self.form_names[0]]['submissions']
        version_uid = self.asset.latest_deployed_version_uid
        export_options = {'hierarchy_in_labels': 'true'}
        expected_lines = [
            '"start";"end";"External Characteristics/What kind of symmetry do you have?";"External Characteristics/What kind of symmetry do you have?/Spherical";"External Characteristics/What kind of symmetry do you have?/Radial";"External Characteristics/What kind of symmetry do you have?/Bilateral";"External Characteristics/How many segments does your body have?";"Do you have body fluids that occupy intracellular space?";"Do you descend from an ancestral unicellular organism?";"_id";"_uuid";"_submission_time";"_validation_status";"_notes";"_status";"_submitted_by";"__version__";"_tags";"_index"',
            '"";"";"#symmetry";"";"";"";"#segments";"#fluids";"";"";"";"";"";"";"";"";"";"";""',
            f'"2017-10-23T05:40:39.000-04:00";"2017-10-23T05:41:13.000-04:00";"Spherical Radial Bilateral";"1";"1";"1";"6";"Yes, and some extracellular space";"No";"{submissions[0]["_id"]}";"48583952-1892-4931-8d9c-869e7b49bafb";"2017-10-23T09:41:19";"";"";"submitted_via_web";"";"{version_uid}";"";"1"',
            f'"2017-10-23T05:41:14.000-04:00";"2017-10-23T05:41:32.000-04:00";"Radial";"0";"1";"0";"3";"Yes";"No";"{submissions[1]["_id"]}";"317ba7b7-bea4-4a8c-8620-a483c3079c4b";"2017-10-23T09:41:38";"";"";"submitted_via_web";"";"{version_uid}";"";"2"',
            f'"2017-10-23T05:41:32.000-04:00";"2017-10-23T05:42:05.000-04:00";"Bilateral";"0";"0";"1";"2";"No / Unsure";"Yes";"{submissions[2]["_id"]}";"3f15cdfe-3eab-4678-8352-7806febf158d";"2017-10-23T09:42:11";"";"";"submitted_via_web";"anotheruser";"{version_uid}";"";"3"',
        ]
        self.run_csv_export_test(expected_lines, export_options)

    def test_csv_export_filter_fields(self):
        export_options = {
            'fields': [
                'start',
                'end',
                'Do_you_descend_from_unicellular_organism',
                '_index',
            ]
        }
        expected_lines = [
            '"start";"end";"Do you descend from an ancestral unicellular organism?";"_uuid";"_index"',
            '"2017-10-23T05:40:39.000-04:00";"2017-10-23T05:41:13.000-04:00";"No";"48583952-1892-4931-8d9c-869e7b49bafb";"1"',
            '"2017-10-23T05:41:14.000-04:00";"2017-10-23T05:41:32.000-04:00";"No";"317ba7b7-bea4-4a8c-8620-a483c3079c4b";"2"',
            '"2017-10-23T05:41:32.000-04:00";"2017-10-23T05:42:05.000-04:00";"Yes";"3f15cdfe-3eab-4678-8352-7806febf158d";"3"',
        ]
        self.run_csv_export_test(expected_lines, export_options)

    def test_xls_export_english_labels(self):
        submissions = self.forms[self.form_names[0]]['submissions']
        version_uid = self.asset.latest_deployed_version_uid
        export_options = {'lang': 'English'}
        expected_data = {
            self.asset.name: [
                [
                    'start',
                    'end',
                    'What kind of symmetry do you have?',
                    'What kind of symmetry do you have?/Spherical',
                    'What kind of symmetry do you have?/Radial',
                    'What kind of symmetry do you have?/Bilateral',
                    'How many segments does your body have?',
                    'Do you have body fluids that occupy intracellular space?',
                    'Do you descend from an ancestral unicellular organism?',
                    '_id',
                    '_uuid',
                    '_submission_time',
                    '_validation_status',
                    '_notes',
                    '_status',
                    '_submitted_by',
                    '__version__',
                    '_tags',
                    '_index',
                ],
                [
                    '',
                    '',
                    '#symmetry',
                    '',
                    '',
                    '',
                    '#segments',
                    '#fluids',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                ],
                [
                    '2017-10-23T05:40:39.000-04:00',
                    '2017-10-23T05:41:13.000-04:00',
                    'Spherical Radial Bilateral',
                    '1',
                    '1',
                    '1',
                    '6',
                    'Yes, and some extracellular space',
                    'No',
                    submissions[0]['_id'],
                    '48583952-1892-4931-8d9c-869e7b49bafb',
                    '2017-10-23T09:41:19',
                    '',
                    '',
                    'submitted_via_web',
                    '',
                    version_uid,
                    '',
                    1.0,
                ],
                [
                    '2017-10-23T05:41:14.000-04:00',
                    '2017-10-23T05:41:32.000-04:00',
                    'Radial',
                    '0',
                    '1',
                    '0',
                    '3',
                    'Yes',
                    'No',
                    submissions[1]['_id'],
                    '317ba7b7-bea4-4a8c-8620-a483c3079c4b',
                    '2017-10-23T09:41:38',
                    '',
                    '',
                    'submitted_via_web',
                    '',
                    version_uid,
                    '',
                    2.0,
                ],
                [
                    '2017-10-23T05:41:32.000-04:00',
                    '2017-10-23T05:42:05.000-04:00',
                    'Bilateral',
                    '0',
                    '0',
                    '1',
                    '2',
                    'No / Unsure',
                    'Yes',
                    submissions[2]['_id'],
                    '3f15cdfe-3eab-4678-8352-7806febf158d',
                    '2017-10-23T09:42:11',
                    '',
                    '',
                    'submitted_via_web',
                    'anotheruser',
                    version_uid,
                    '',
                    3.0,
                ],
            ]
        }
        self.run_xls_export_test(expected_data, export_options)

    def test_xls_export_english_labels_partial_submissions(self):
        submissions = self.forms[self.form_names[0]]['submissions']
        version_uid = self.asset.latest_deployed_version_uid
        export_options = {'lang': 'English'}
        expected_data = {
            self.asset.name: [
                [
                    'start',
                    'end',
                    'What kind of symmetry do you have?',
                    'What kind of symmetry do you have?/Spherical',
                    'What kind of symmetry do you have?/Radial',
                    'What kind of symmetry do you have?/Bilateral',
                    'How many segments does your body have?',
                    'Do you have body fluids that occupy intracellular space?',
                    'Do you descend from an ancestral unicellular organism?',
                    '_id',
                    '_uuid',
                    '_submission_time',
                    '_validation_status',
                    '_notes',
                    '_status',
                    '_submitted_by',
                    '__version__',
                    '_tags',
                    '_index',
                ],
                [
                    '',
                    '',
                    '#symmetry',
                    '',
                    '',
                    '',
                    '#segments',
                    '#fluids',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                ],
                [
                    '2017-10-23T05:41:32.000-04:00',
                    '2017-10-23T05:42:05.000-04:00',
                    'Bilateral',
                    '0',
                    '0',
                    '1',
                    '2',
                    'No / Unsure',
                    'Yes',
                    submissions[2]['_id'],
                    '3f15cdfe-3eab-4678-8352-7806febf158d',
                    '2017-10-23T09:42:11',
                    '',
                    '',
                    'submitted_via_web',
                    'anotheruser',
                    version_uid,
                    '',
                    1.0,
                ],
            ]
        }
        self.run_xls_export_test(expected_data, export_options, user=self.anotheruser)

    def test_xls_export_multiple_select_both(self):
        submissions = self.forms[self.form_names[0]]['submissions']
        version_uid = self.asset.latest_deployed_version_uid
        export_options = {'lang': 'English', 'multiple_select': 'both'}
        expected_data = {
            self.asset.name: [
                [
                    'start',
                    'end',
                    'What kind of symmetry do you have?',
                    'What kind of symmetry do you have?/Spherical',
                    'What kind of symmetry do you have?/Radial',
                    'What kind of symmetry do you have?/Bilateral',
                    'How many segments does your body have?',
                    'Do you have body fluids that occupy intracellular space?',
                    'Do you descend from an ancestral unicellular organism?',
                    '_id',
                    '_uuid',
                    '_submission_time',
                    '_validation_status',
                    '_notes',
                    '_status',
                    '_submitted_by',
                    '__version__',
                    '_tags',
                    '_index',
                ],
                [
                    '',
                    '',
                    '#symmetry',
                    '',
                    '',
                    '',
                    '#segments',
                    '#fluids',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                ],
                [
                    '2017-10-23T05:40:39.000-04:00',
                    '2017-10-23T05:41:13.000-04:00',
                    'Spherical Radial Bilateral',
                    '1',
                    '1',
                    '1',
                    '6',
                    'Yes, and some extracellular space',
                    'No',
                    submissions[0]['_id'],
                    '48583952-1892-4931-8d9c-869e7b49bafb',
                    '2017-10-23T09:41:19',
                    '',
                    '',
                    'submitted_via_web',
                    '',
                    version_uid,
                    '',
                    1.0,
                ],
                [
                    '2017-10-23T05:41:14.000-04:00',
                    '2017-10-23T05:41:32.000-04:00',
                    'Radial',
                    '0',
                    '1',
                    '0',
                    '3',
                    'Yes',
                    'No',
                    submissions[1]['_id'],
                    '317ba7b7-bea4-4a8c-8620-a483c3079c4b',
                    '2017-10-23T09:41:38',
                    '',
                    '',
                    'submitted_via_web',
                    '',
                    version_uid,
                    '',
                    2.0,
                ],
                [
                    '2017-10-23T05:41:32.000-04:00',
                    '2017-10-23T05:42:05.000-04:00',
                    'Bilateral',
                    '0',
                    '0',
                    '1',
                    '2',
                    'No / Unsure',
                    'Yes',
                    submissions[2]['_id'],
                    '3f15cdfe-3eab-4678-8352-7806febf158d',
                    '2017-10-23T09:42:11',
                    '',
                    '',
                    'submitted_via_web',
                    'anotheruser',
                    version_uid,
                    '',
                    3.0,
                ],
            ]
        }
        self.run_xls_export_test(expected_data, export_options)

    def test_xls_export_multiple_select_summary(self):
        submissions = self.forms[self.form_names[0]]['submissions']
        version_uid = self.asset.latest_deployed_version_uid
        export_options = {'lang': 'English', 'multiple_select': 'summary'}
        expected_data = {
            self.asset.name: [
                [
                    'start',
                    'end',
                    'What kind of symmetry do you have?',
                    'How many segments does your body have?',
                    'Do you have body fluids that occupy intracellular space?',
                    'Do you descend from an ancestral unicellular organism?',
                    '_id',
                    '_uuid',
                    '_submission_time',
                    '_validation_status',
                    '_notes',
                    '_status',
                    '_submitted_by',
                    '__version__',
                    '_tags',
                    '_index',
                ],
                [
                    '',
                    '',
                    '#symmetry',
                    '#segments',
                    '#fluids',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                ],
                [
                    '2017-10-23T05:40:39.000-04:00',
                    '2017-10-23T05:41:13.000-04:00',
                    'Spherical Radial Bilateral',
                    '6',
                    'Yes, and some extracellular space',
                    'No',
                    submissions[0]['_id'],
                    '48583952-1892-4931-8d9c-869e7b49bafb',
                    '2017-10-23T09:41:19',
                    '',
                    '',
                    'submitted_via_web',
                    '',
                    version_uid,
                    '',
                    1.0,
                ],
                [
                    '2017-10-23T05:41:14.000-04:00',
                    '2017-10-23T05:41:32.000-04:00',
                    'Radial',
                    '3',
                    'Yes',
                    'No',
                    submissions[1]['_id'],
                    '317ba7b7-bea4-4a8c-8620-a483c3079c4b',
                    '2017-10-23T09:41:38',
                    '',
                    '',
                    'submitted_via_web',
                    '',
                    version_uid,
                    '',
                    2.0,
                ],
                [
                    '2017-10-23T05:41:32.000-04:00',
                    '2017-10-23T05:42:05.000-04:00',
                    'Bilateral',
                    '2',
                    'No / Unsure',
                    'Yes',
                    submissions[2]['_id'],
                    '3f15cdfe-3eab-4678-8352-7806febf158d',
                    '2017-10-23T09:42:11',
                    '',
                    '',
                    'submitted_via_web',
                    'anotheruser',
                    version_uid,
                    '',
                    3.0,
                ],
            ]
        }
        self.run_xls_export_test(expected_data, export_options)

    def test_xls_export_multiple_select_details(self):
        submissions = self.forms[self.form_names[0]]['submissions']
        version_uid = self.asset.latest_deployed_version_uid
        export_options = {'lang': 'English', 'multiple_select': 'details'}
        expected_data = {
            self.asset.name: [
                [
                    'start',
                    'end',
                    'What kind of symmetry do you have?/Spherical',
                    'What kind of symmetry do you have?/Radial',
                    'What kind of symmetry do you have?/Bilateral',
                    'How many segments does your body have?',
                    'Do you have body fluids that occupy intracellular space?',
                    'Do you descend from an ancestral unicellular organism?',
                    '_id',
                    '_uuid',
                    '_submission_time',
                    '_validation_status',
                    '_notes',
                    '_status',
                    '_submitted_by',
                    '__version__',
                    '_tags',
                    '_index',
                ],
                [
                    '',
                    '',
                    '#symmetry',
                    '',
                    '',
                    '#segments',
                    '#fluids',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                ],
                [
                    '2017-10-23T05:40:39.000-04:00',
                    '2017-10-23T05:41:13.000-04:00',
                    '1',
                    '1',
                    '1',
                    '6',
                    'Yes, and some extracellular space',
                    'No',
                    submissions[0]['_id'],
                    '48583952-1892-4931-8d9c-869e7b49bafb',
                    '2017-10-23T09:41:19',
                    '',
                    '',
                    'submitted_via_web',
                    '',
                    version_uid,
                    '',
                    1.0,
                ],
                [
                    '2017-10-23T05:41:14.000-04:00',
                    '2017-10-23T05:41:32.000-04:00',
                    '0',
                    '1',
                    '0',
                    '3',
                    'Yes',
                    'No',
                    submissions[1]['_id'],
                    '317ba7b7-bea4-4a8c-8620-a483c3079c4b',
                    '2017-10-23T09:41:38',
                    '',
                    '',
                    'submitted_via_web',
                    '',
                    version_uid,
                    '',
                    2.0,
                ],
                [
                    '2017-10-23T05:41:32.000-04:00',
                    '2017-10-23T05:42:05.000-04:00',
                    '0',
                    '0',
                    '1',
                    '2',
                    'No / Unsure',
                    'Yes',
                    submissions[2]['_id'],
                    '3f15cdfe-3eab-4678-8352-7806febf158d',
                    '2017-10-23T09:42:11',
                    '',
                    '',
                    'submitted_via_web',
                    'anotheruser',
                    version_uid,
                    '',
                    3.0,
                ],
            ]
        }
        self.run_xls_export_test(expected_data, export_options)

    def test_xls_export_filter_fields(self):
        export_options = {
            'fields': [
                'start',
                'end',
                'Do_you_descend_from_unicellular_organism',
                '_index',
            ]
        }
        expected_data = {
            self.asset.name: [
                [
                    'start',
                    'end',
                    'Do you descend from an ancestral unicellular organism?',
                    '_uuid',
                    '_index',
                ],
                [
                    '2017-10-23T05:40:39.000-04:00',
                    '2017-10-23T05:41:13.000-04:00',
                    'No',
                    '48583952-1892-4931-8d9c-869e7b49bafb',
                    1.0,
                ],
                [
                    '2017-10-23T05:41:14.000-04:00',
                    '2017-10-23T05:41:32.000-04:00',
                    'No',
                    '317ba7b7-bea4-4a8c-8620-a483c3079c4b',
                    2.0,
                ],
                [
                    '2017-10-23T05:41:32.000-04:00',
                    '2017-10-23T05:42:05.000-04:00',
                    'Yes',
                    '3f15cdfe-3eab-4678-8352-7806febf158d',
                    3.0,
                ],
            ]
        }
        self.run_xls_export_test(expected_data, export_options)

    def test_xls_export_filter_fields_without_index(self):
        export_options = {
            'fields': [
                'start',
                'end',
                'Do_you_descend_from_unicellular_organism',
            ]
        }
        expected_data = {
            self.asset.name: [
                [
                    'start',
                    'end',
                    'Do you descend from an ancestral unicellular organism?',
                    '_uuid',
                ],
                [
                    '2017-10-23T05:40:39.000-04:00',
                    '2017-10-23T05:41:13.000-04:00',
                    'No',
                    '48583952-1892-4931-8d9c-869e7b49bafb',
                ],
                [
                    '2017-10-23T05:41:14.000-04:00',
                    '2017-10-23T05:41:32.000-04:00',
                    'No',
                    '317ba7b7-bea4-4a8c-8620-a483c3079c4b',
                ],
                [
                    '2017-10-23T05:41:32.000-04:00',
                    '2017-10-23T05:42:05.000-04:00',
                    'Yes',
                    '3f15cdfe-3eab-4678-8352-7806febf158d',
                ],
            ]
        }
        self.run_xls_export_test(expected_data, export_options)

    def test_xls_export_filter_fields_with_media_url(self):
        asset_name = 'Simple media'
        export_options = {'fields': ['an_image'], 'include_media_url': True}
        asset = self.assets[asset_name]
        submissions = self.forms[asset_name]['submissions']
        submission = asset.deployment.get_submission(submissions[0]['_id'], asset.owner)
        media_url = submission['_attachments'][0]['download_url']
        expected_data = {
            asset_name: [
                ['Submit an image', 'Submit an image_URL', '_uuid'],
                [
                    'audio_conversion_test_image.jpg',
                    media_url,
                    'f80be949-89b5-4af1-a42d-7d292b2bc0cd',
                ],
            ]
        }
        self.run_xls_export_test(
            expected_data, export_options, asset=self.assets[asset_name]
        )

    def test_xls_export_filter_fields_repeat_groups(self):
        export_options = {
            'fields': [
                '_uuid',
                '_submission_time',
                'people/person/name',
                '_index'
            ]
        }
        asset = self.assets['Simple repeat group']
        expected_data = {
            asset.name: [
                [
                    '_uuid',
                    '_submission_time',
                    '_index',
                ],
                [
                    'f80be949-89b5-4af1-a29d-7d292b2bc0cd',
                    '2021-06-30T22:12:56',
                    1.0,
                ],
            ],
            'person': [
                [
                    'name',
                    '_index',
                    '_parent_table_name',
                    '_parent_index',
                    '_submission__id',
                    '_submission__uuid',
                    '_submission__submission_time',
                    '_submission__validation_status',
                    '_submission__notes',
                    '_submission__status',
                    '_submission__submitted_by',
                    '_submission___version__',
                    '_submission__tags',
                ],
                [
                    'Julius Caesar',
                    1.0,
                    'Simple repeat group',
                    1.0,
                    '',
                    'f80be949-89b5-4af1-a29d-7d292b2bc0cd',
                    '2021-06-30T22:12:56',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                ],
                [
                    'Augustus',
                    2.0,
                    'Simple repeat group',
                    1.0,
                    '',
                    'f80be949-89b5-4af1-a29d-7d292b2bc0cd',
                    '2021-06-30T22:12:56',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                ],
            ],
        }
        self.run_xls_export_test(
            expected_data,
            export_options,
            asset=asset,
            repeat_group=True,
        )

    def test_xls_export_repeat_groups(self):
        asset = self.assets['Simple repeat group']
        submissions = self.forms['Simple repeat group']['submissions']
        version_uid = asset.latest_deployed_version_uid
        expected_data = {
            asset.name: [
                [
                    '_id',
                    '_uuid',
                    '_submission_time',
                    '_validation_status',
                    '_notes',
                    '_status',
                    '_submitted_by',
                    '__version__',
                    '_tags',
                    '_index',
                ],
                [
                    submissions[0]['_id'],
                    'f80be949-89b5-4af1-a29d-7d292b2bc0cd',
                    '2021-06-30T22:12:56',
                    '',
                    '',
                    'submitted_via_web',
                    '',
                    version_uid,
                    '',
                    1.0,
                ],
            ],
            'person': [
                [
                    'name',
                    'age',
                    '_index',
                    '_parent_table_name',
                    '_parent_index',
                    '_submission__id',
                    '_submission__uuid',
                    '_submission__submission_time',
                    '_submission__validation_status',
                    '_submission__notes',
                    '_submission__status',
                    '_submission__submitted_by',
                    '_submission___version__',
                    '_submission__tags',
                ],
                [
                    'Julius Caesar',
                    '55',
                    1.0,
                    'Simple repeat group',
                    1.0,
                    submissions[0]['_id'],
                    'f80be949-89b5-4af1-a29d-7d292b2bc0cd',
                    '2021-06-30T22:12:56',
                    '',
                    '',
                    'submitted_via_web',
                    '',
                    version_uid,
                    '',
                ],
                [
                    'Augustus',
                    '75',
                    2.0,
                    'Simple repeat group',
                    1.0,
                    submissions[0]['_id'],
                    'f80be949-89b5-4af1-a29d-7d292b2bc0cd',
                    '2021-06-30T22:12:56',
                    '',
                    '',
                    'submitted_via_web',
                    '',
                    version_uid,
                    '',
                ],
            ],
        }
        self.run_xls_export_test(expected_data, asset=asset, repeat_group=True)

    def test_export_spss_labels(self):
        export_task = SubmissionExportTask()
        export_task.user = self.user
        export_task.data = {
            'source': reverse('asset-detail', args=[self.asset.uid]),
            'type': 'spss_labels',
        }
        messages = defaultdict(list)
        # Set the current date and time artificially to generate a predictable
        # file name for the export
        utcnow = datetime.datetime.now(tz=ZoneInfo('UTC'))
        with mock.patch('kpi.models.import_export_task.utcnow') as mock_utcnow:
            mock_utcnow.return_value = utcnow
            export_task._run_task(messages)
        self.assertFalse(messages)
        self.assertEqual(
            os.path.split(export_task.result.name)[-1],
            'Identificaci\xf3n_de_animales_-_all_versions_-_SPSS_Labels_-_'
            '{date:%Y-%m-%d-%H-%M-%S}.zip'.format(date=utcnow)
        )
        expected_file_names_and_content_lines = {
            'Identificaci\xf3n de animales - Spanish - SPSS labels.sps': [
                '\ufeffVARIABLE LABELS',
                " start 'start'",
                " /end 'end'",
                " /What_kind_of_symmetry_do_you_have '\xbfQu\xe9 tipo de simetr\xeda tiene?'",
                " /What_kind_of_symmetry_do_you_have_spherical '\xbfQu\xe9 tipo de simetr\xeda tiene? :: Esf\xe9rico'",
                " /What_kind_of_symmetry_do_you_have_radial '\xbfQu\xe9 tipo de simetr\xeda tiene? :: Radial'",
                " /What_kind_of_symmetry_do_you_have_bilateral '\xbfQu\xe9 tipo de simetr\xeda tiene? :: Bilateral'",
                " /How_many_segments_does_your_body_have '\xbfCu\xe1ntos segmentos tiene tu cuerpo?'",
                " /Do_you_have_body_flu_intracellular_space '\xbfTienes fluidos corporales que ocupan espacio intracelular?'",
                " /Do_you_descend_from_unicellular_organism '\xbfDesciende de un organismo unicelular ancestral?'",
                " /_id '_id'",
                " /_uuid '_uuid'",
                " /_submission_time '_submission_time'",
                " /_validation_status '_validation_status'",
                " /_notes '_notes'",
                " /_status '_status'",
                " /_submitted_by '_submitted_by'",
                " /__version__ '__version__'",
                " /_tags '_tags'",
                ' .',
                'VALUE LABELS',
                ' Do_you_have_body_flu_intracellular_space',
                " 'yes' 'S\xed'",
                " 'yes__and_some_' 'S\xed, y alg\xfan espacio extracelular'",
                " 'no___unsure' 'No / Inseguro'",
                ' /Do_you_descend_from_unicellular_organism',
                " 'yes' 'S\xed'",
                " 'no' 'No'",
                ' .'
            ],
            'Identificaci\xf3n de animales - English - SPSS labels.sps': [
                '\ufeffVARIABLE LABELS',
                " start 'start'",
                " /end 'end'",
                " /What_kind_of_symmetry_do_you_have 'What kind of symmetry do you have?'",
                " /What_kind_of_symmetry_do_you_have_spherical 'What kind of symmetry do you have? :: Spherical'",
                " /What_kind_of_symmetry_do_you_have_radial 'What kind of symmetry do you have? :: Radial'",
                " /What_kind_of_symmetry_do_you_have_bilateral 'What kind of symmetry do you have? :: Bilateral'",
                " /How_many_segments_does_your_body_have 'How many segments does your body have?'",
                " /Do_you_have_body_flu_intracellular_space 'Do you have body fluids that occupy intracellular space?'",
                " /Do_you_descend_from_unicellular_organism 'Do you descend from an ancestral unicellular organism?'",
                " /_id '_id'",
                " /_uuid '_uuid'",
                " /_submission_time '_submission_time'",
                " /_validation_status '_validation_status'",
                " /_notes '_notes'",
                " /_status '_status'",
                " /_submitted_by '_submitted_by'",
                " /__version__ '__version__'",
                " /_tags '_tags'",
                ' .',
                'VALUE LABELS',
                ' Do_you_have_body_flu_intracellular_space',
                " 'yes' 'Yes'",
                " 'yes__and_some_' 'Yes, and some extracellular space'",
                " 'no___unsure' 'No / Unsure'",
                ' /Do_you_descend_from_unicellular_organism',
                " 'yes' 'Yes'",
                " 'no' 'No'",
                ' .'
            ],
        }
        result_zip = zipfile.ZipFile(export_task.result, 'r')
        for name, content_lines in expected_file_names_and_content_lines.items():
            self.assertEqual(
                # we have `unicode_literals` but the rest of the app doesn't
                result_zip.open(name, 'r').read().decode('utf-8'),
                '\r\n'.join(content_lines)
            )

    def test_remove_excess_exports(self):
        task_data = {
            'source': reverse('asset-detail', args=[self.asset.uid]),
            'type': 'csv',
        }
        # Create and run one export, so we can verify that it's `result` file
        # is later deleted
        export_task = SubmissionExportTask()
        export_task.user = self.user
        export_task.data = task_data
        export_task.save()
        export_task.run()
        self.assertEqual(export_task.status, SubmissionExportTask.COMPLETE)
        result = export_task.result
        self.assertTrue(result.storage.exists(result.name))
        # Make an excessive amount of additional exports
        excess_count = 5 + settings.MAXIMUM_EXPORTS_PER_USER_PER_FORM
        for _ in range(excess_count):
            export_task = SubmissionExportTask()
            export_task.user = self.user
            export_task.data = task_data
            export_task.save()
        created_export_tasks = SubmissionExportTask.objects.filter(
            user=self.user, data__source=task_data['source']
        )
        self.assertEqual(excess_count + 1, created_export_tasks.count())
        # Identify which exports should be kept
        export_tasks_to_keep = created_export_tasks.order_by('-date_created')[
            :settings.MAXIMUM_EXPORTS_PER_USER_PER_FORM]
        # Call `run()` once more since it invokes the cleanup logic
        export_task.run()
        self.assertEqual(export_task.status, SubmissionExportTask.COMPLETE)
        # Verify the cleanup
        self.assertFalse(result.storage.exists(result.name))
        self.assertListEqual(  # assertSequenceEqual isn't working...
            list(export_tasks_to_keep.values_list('pk', flat=True)),
            list(
                SubmissionExportTask.objects.filter(
                    user=self.user, data__source=task_data['source']
                ).order_by('-date_created').values_list('pk', flat=True)
            ),
        )

    def test_log_and_mark_stuck_exports_as_errored(self):
        task_data = {
            'source': reverse('asset-detail', args=[self.asset.uid]),
            'type': 'csv',
        }
        self.assertEqual(
            0,
            SubmissionExportTask.objects.filter(
                user=self.user, data__source=task_data['source']
            ).count(),
        )
        # Simulate a few stuck exports
        for status in (SubmissionExportTask.CREATED, SubmissionExportTask.PROCESSING):
            export_task = SubmissionExportTask()
            export_task.user = self.user
            export_task.data = task_data
            export_task.status = status
            export_task.save()
            export_task.date_created -= datetime.timedelta(days=1)
            export_task.save()
        self.assertSequenceEqual(
            [SubmissionExportTask.CREATED, SubmissionExportTask.PROCESSING],
            SubmissionExportTask.objects.filter(
                user=self.user, data__source=task_data['source']
            ).order_by('pk').values_list('status', flat=True),
        )
        # Run another export, which invokes the cleanup logic
        export_task = SubmissionExportTask()
        export_task.user = self.user
        export_task.data = task_data
        export_task.save()
        export_task.run()
        # Verify that the stuck exports have been marked
        self.assertSequenceEqual(
            [SubmissionExportTask.ERROR, SubmissionExportTask.ERROR, SubmissionExportTask.COMPLETE],
            SubmissionExportTask.objects.filter(
                user=self.user, data__source=task_data['source']
            ).order_by('pk').values_list('status', flat=True),
        )

    def test_export_long_form_title(self):
        what_a_title = (
            'the quick brown fox jumped over the lazy dog and jackdaws love '
            'my big sphinx of quartz and pack my box with five dozen liquor '
            'jugs dum cornelia legit flavia scribit et laeta est flavia quod '
            'cornelia iam in villa habitat et cornelia et flavia sunt amicae'
        )
        assert len(what_a_title) > SubmissionExportTask.MAXIMUM_FILENAME_LENGTH
        self.asset.name = what_a_title
        self.asset.save()
        task_data = {
            'source': reverse('asset-detail', args=[self.asset.uid]),
            'type': 'csv',
        }
        export_task = SubmissionExportTask()
        export_task.user = self.user
        export_task.data = task_data
        export_task.save()
        export_task.run()

        assert (
            len(os.path.basename(export_task.result.name)) ==
                SubmissionExportTask.MAXIMUM_FILENAME_LENGTH
        )

    def test_export_latest_version_only(self):
        submissions = self.forms[self.form_names[0]]['submissions']
        new_survey_content = [{
            'label': ['Do you descend... new label',
                      '\xbfDesciende de... etiqueta nueva'],
            'name': 'Do_you_descend_from_unicellular_organism',
            'required': False,
            'type': 'text'
        }]
        # Re-fetch from the database to avoid modifying self.form_content
        self.asset = Asset.objects.get(pk=self.asset.pk)
        self.asset.content['survey'] = new_survey_content
        self.asset.save()
        # Retrieve the uid of last deployed version before deploying. Otherwise,
        # it won't match with export data because data has been submitted prior
        # this deployment.
        version_uid = self.asset.latest_deployed_version_uid
        self.asset.deploy(backend='mock', active=True)
        expected_lines = [
            '"Do you descend... new label";"_id";"_uuid";"_submission_time";"_validation_status";"_notes";"_status";"_submitted_by";"__version__";"_tags";"_index"',
            f'"no";"{submissions[0]["_id"]}";"48583952-1892-4931-8d9c-869e7b49bafb";"2017-10-23T09:41:19";"";"";"submitted_via_web";"";"{version_uid}";"";"1"',
            f'"no";"{submissions[1]["_id"]}";"317ba7b7-bea4-4a8c-8620-a483c3079c4b";"2017-10-23T09:41:38";"";"";"submitted_via_web";"";"{version_uid}";"";"2"',
            f'"yes";"{submissions[2]["_id"]}";"3f15cdfe-3eab-4678-8352-7806febf158d";"2017-10-23T09:42:11";"";"";"submitted_via_web";"anotheruser";"{version_uid}";"";"3"',
        ]
        self.run_csv_export_test(
            expected_lines, {'fields_from_all_versions': 'false'})

    def test_export_exceeding_api_submission_limit(self):
        """
        Make sure the limit on count of submissions returned by the API does
        not apply to exports
        """
        limit = settings.SUBMISSION_LIST_LIMIT
        excess = 10
        asset = Asset.objects.create(
            name='Lots of submissions',
            owner=self.user,
            content={'survey': [{'label': 'q', 'name': 'q', 'type': 'integer'}]},
        )
        asset.deploy(backend='mock', active=True)
        submissions = [
            {
                '__version__': asset.latest_deployed_version.uid,
                'q': i,
            } for i in range(limit + excess)
        ]
        asset.deployment.mock_submissions(submissions)
        export_task = SubmissionExportTask()
        export_task.user = self.user
        export_task.data = {
            'source': reverse('asset-detail', args=[asset.uid]),
            'type': 'csv'
        }
        messages = defaultdict(list)
        export_task._run_task(messages)
        # Don't forget to add one for the header row!
        self.assertEqual(len(list(export_task.result)), limit + excess + 1)

    def test_export_with_disabled_questions(self):
        asset = Asset.objects.create(
            name='Form with undocumented `disabled` column',
            owner=self.user,
            content={
                'survey': [
                    {'label': 'q', 'name': 'q', 'type': 'integer'},
                    {'name': 'ignore', 'type': 'select_one nope', 'disabled': True},
                ]
            },
        )
        asset.deploy(backend='mock', active=True)
        submissions = [
            {
                '__version__': asset.latest_deployed_version.uid,
                'q': 123,
                '_submission_time': '2017-10-23T09:41:19',
            }
        ]
        asset.deployment.mock_submissions(submissions)
        # observe that `ignore` does not appear!
        expected_lines = [
            '"q";"_id";"_uuid";"_submission_time";"_validation_status";"_notes";"_status";"_submitted_by";"__version__";"_tags";"_index"',
            f'"123";"{submissions[0]["_id"]}";"{submissions[0]["_uuid"]}";"2017-10-23T09:41:19";"";"";"submitted_via_web";"someuser";"{asset.latest_deployed_version.uid}";"";"1"',
        ]
        # fails with `KeyError` prior to fix for kobotoolbox/formpack#219
        self.run_csv_export_test(expected_lines, asset=asset)

    def test_anotheruser_can_export_when_submissions_publicly_shared(self):
        """
        Running through behaviour described in issue kpi/#2870 where an asset
        that has been publicly shared and then explicity shared with a user, the
        user has lower permissions than an anonymous user and is therefore
        unable to export submission data.
        """
        # resetting permissions of `anotheruser` to have no permissions
        self.asset.remove_perm(self.anotheruser, PERM_PARTIAL_SUBMISSIONS)
        self.asset.remove_perm(self.anotheruser, PERM_VIEW_ASSET)

        anonymous_user = get_anonymous_user()

        assert self.asset.has_perm(self.anotheruser, PERM_VIEW_ASSET) == False
        assert PERM_VIEW_ASSET not in self.asset.get_perms(self.anotheruser)
        assert self.asset.has_perm(self.anotheruser, PERM_CHANGE_ASSET) == False
        assert PERM_CHANGE_ASSET not in self.asset.get_perms(self.anotheruser)

        # required to export
        self.asset.assign_perm(self.anotheruser, PERM_CHANGE_ASSET)

        assert self.asset.has_perm(self.anotheruser, PERM_VIEW_ASSET) == True
        assert PERM_VIEW_ASSET in self.asset.get_perms(self.anotheruser)
        assert self.asset.has_perm(self.anotheruser, PERM_CHANGE_ASSET) == True
        assert PERM_CHANGE_ASSET in self.asset.get_perms(self.anotheruser)

        assert (
            self.asset.has_perm(self.anotheruser, PERM_VIEW_SUBMISSIONS)
            == False
        )
        assert PERM_VIEW_SUBMISSIONS not in self.asset.get_perms(
            self.anotheruser
        )

        self.asset.assign_perm(anonymous_user, PERM_VIEW_SUBMISSIONS)

        assert (
            self.asset.has_perm(self.anotheruser, PERM_VIEW_SUBMISSIONS) == True
        )
        assert PERM_VIEW_SUBMISSIONS in self.asset.get_perms(self.anotheruser)

        # testing anotheruser can export data
        self.run_csv_export_test(user=self.anotheruser)
